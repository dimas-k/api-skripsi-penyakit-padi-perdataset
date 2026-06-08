"""
isi_excel_untuk_validator.py  (v3 — Penilaian Ground Truth)
=============================================================
Perubahan dari v2:
  - HAPUS kolom Jawaban Sistem AI (3 LLM)
  - HAPUS kolom Skor LOW / MEDIUM / HIGH
  - TAMBAH 2 kolom skor validator: Skor GT-1 dan Skor GT-2
  - Validator (petani) menilai kelengkapan & ketepatan masing-masing Ground Truth
  - Layout: A4 Landscape, 7 kolom
    No | Penyakit | Query | GT-1 | GT-2 | Skor GT-1 | Skor GT-2

Cara pakai:
  python isi_excel_untuk_validator.py
  python isi_excel_untuk_validator.py --input hasil_evaluasi_rag.json --kb knowledge_base/penyakit_padi.txt
"""

import json, re, argparse
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

# ─── Konfigurasi ────────────────────────────────────────────────
JSON_PATH  = "hasil_evaluasi_rag.json"
KB_PATH    = "knowledge_base/penyakit_padi.txt"
OUTPUT     = "Penilaian_Faithfulness_untuk_Validator.xlsx"

MAX_CHARS_GT = 1000  # max karakter ground truth per kolom

TIER_KEYWORDS = {
    "LOW"   : ["LOW",  "low",  "qwen", "Qwen", "3b", "3B"],
    "MEDIUM": ["MEDIUM", "medium", "Gemini", "gemini", "Flash"],
    "HIGH"  : ["HIGH", "high", "llama", "Llama", "70b", "70B",
               "versatile", "groq", "Groq"],
}


# ═════════════════════════════════════════════════════════════════
# PARSE KNOWLEDGE BASE → Ground Truth ke-2
# ═════════════════════════════════════════════════════════════════
def parse_kb(kb_path: str) -> dict[str, str]:
    """
    Baca penyakit_padi.txt dan ekstrak GEJALA + PENANGANAN per kode penyakit.
    Kembalikan dict: {kode_penyakit: teks_gt2}
    """
    text = Path(kb_path).read_text(encoding="utf-8")
    bloks = re.split(r"={3,}", text)

    result = {}
    kode    = None
    nama_id = None

    for blok in bloks:
        m = re.search(r"Kode:\s*(\w+)", blok)
        if m:
            kode = m.group(1).strip().lower()

        mn = re.search(r"(?:Nama(?:\s+Penyakit)?|Penyakit)\s*:\s*(.+)", blok)
        if mn:
            nama_id = mn.group(1).strip()

        if not kode:
            continue

        # Ekstrak GEJALA
        gejala = ""
        mg = re.search(r"GEJALA\s*:\s*(.*?)(?=\n[A-Z]{3,}|\Z)", blok, re.S)
        if mg:
            gejala = mg.group(1).strip()
        if not gejala:
            mg = re.search(r"CIRI-CIRI[^:]*:\s*(.*?)(?=\n[A-Z]{3,}|\Z)", blok, re.S)
            if mg:
                gejala = "Ciri-ciri: " + mg.group(1).strip()
        if not gejala:
            mg = re.search(r"INDIKATOR[^:]*:\s*(.*?)(?=\n[A-Z]{3,}|\Z)", blok, re.S)
            if mg:
                gejala = "Indikator: " + mg.group(1).strip()
        if gejala:
            gejala = re.sub(r"\s+", " ", gejala)
            if len(gejala) > 500:
                gejala = gejala[:500].rsplit(" ", 1)[0] + "..."

        # Ekstrak PENANGANAN
        penanganan = ""
        mp = re.search(
            r"PENANG[A-Z]+(?:\s+[A-Z]+)*\s*:\s*(.*?)(?=\n[A-Z]{4,}|\Z)",
            blok, re.S
        )
        if not mp:
            mp = re.search(r"PEMELIHARAAN[^:]*:\s*(.*?)(?=\n[A-Z]{4,}|\Z)", blok, re.S)
        if mp:
            raw = mp.group(1).strip()
            baris = re.findall(r"\d+\.\s+[^\n]+", raw)
            penanganan = " | ".join(b.strip() for b in baris[:3])
            if len(penanganan) > 500:
                penanganan = penanganan[:500].rsplit(" ", 1)[0] + "..."

        if gejala or penanganan:
            header = f"{nama_id} ({kode})" if nama_id else kode
            gt2 = f"[{header}]"
            if gejala:
                gt2 += f"\nGejala: {gejala}"
            if penanganan:
                gt2 += f"\nPenanganan: {penanganan}"
            result[kode] = gt2.strip()
            kode    = None
            nama_id = None

    return result


# ═════════════════════════════════════════════════════════════════
# UTILS
# ═════════════════════════════════════════════════════════════════
def load_json(path: str) -> dict:
    return json.loads(Path(path).read_text(encoding="utf-8"))


def extract_generators(d: dict) -> list[dict]:
    raw = []
    if "generators" in d and isinstance(d["generators"], list):
        raw = [g for g in d["generators"] if "error" not in g]
    elif "generator" in d and isinstance(d["generator"], dict):
        raw = [g for g in d["generator"].values()
               if isinstance(g, dict) and "error" not in g]

    def tier_order(g):
        name = g.get("llm", "")
        for i, (tier, kws) in enumerate(TIER_KEYWORDS.items()):
            if any(kw in name for kw in kws):
                return i
        return 99

    return sorted(raw, key=tier_order)


# ═════════════════════════════════════════════════════════════════
# STYLING
# ═════════════════════════════════════════════════════════════════
def make_border(style_type="thin"):
    s = Side(style=style_type)
    return Border(left=s, right=s, top=s, bottom=s)


def cell_style(cell, bold=False, fill_hex=None, size=12, wrap=True,
               halign="left", valign="top", color="000000"):
    cell.font      = Font(bold=bold, size=size, name="Arial Narrow", color=color)
    cell.alignment = Alignment(horizontal=halign, vertical=valign,
                               wrap_text=wrap, shrink_to_fit=False)
    cell.border    = make_border()
    if fill_hex:
        cell.fill = PatternFill("solid", fgColor=fill_hex)


# ═════════════════════════════════════════════════════════════════
# BUILD EXCEL
# ═════════════════════════════════════════════════════════════════
def build_excel(d: dict, kb_data: dict, output_path: str):
    generators = extract_generators(d)
    if not generators:
        print("❌ Tidak ada data generator di JSON.")
        return

    first_gen = generators[0]["per_query"]

    wb = Workbook()

    # ════════════════════════════════════════════════════════════
    # SHEET 1: Penilaian Faithfulness
    # ════════════════════════════════════════════════════════════
    ws = wb.active
    ws.title = "Penilaian Faithfulness"

    # Setup halaman A4 Landscape
    ws.page_setup.orientation    = "landscape"
    ws.page_setup.paperSize      = 9
    ws.page_setup.fitToPage      = True
    ws.page_setup.fitToWidth     = 1
    ws.page_setup.fitToHeight    = 0
    ws.page_setup.scale          = 100
    ws.print_options.horizontalCentered = True
    ws.page_margins = PageMargins(
        left=0.4, right=0.4, top=0.5, bottom=0.5,
        header=0.2, footer=0.2
    )
    ws.oddHeader.center.text = "LEMBAR PENILAIAN FAITHFULNESS — Sistem RAG Penyakit Padi"
    ws.oddFooter.right.text  = "Halaman &P dari &N"

    # ── Struktur kolom (7 kolom total, tanpa LLM) ───────────────
    # A=No | B=Penyakit | C=Query | D=GT-1 | E=GT-2 | F=Skor GT-1 | G=Skor GT-2
    COL_NO      = 1
    COL_PYK     = 2
    COL_QRY     = 3
    COL_GT1     = 4
    COL_GT2     = 5
    COL_SKOR_GT1 = 6   # validator isi: seberapa baik GT-1
    COL_SKOR_GT2 = 7   # validator isi: seberapa baik GT-2
    TOTAL_COLS  = 7

    col_widths = {
        COL_NO       : 4,
        COL_PYK      : 20,
        COL_QRY      : 32,
        COL_GT1      : 55,
        COL_GT2      : 55,
        COL_SKOR_GT1 : 14,
        COL_SKOR_GT2 : 14,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    last_col = get_column_letter(TOTAL_COLS)

    # ── Baris 1: Judul ───────────────────────────────────────────
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "LEMBAR PENILAIAN FAITHFULNESS — Sistem Rekomendasi Penyakit Padi"
    cell_style(ws["A1"], bold=True, fill_hex="1B5E20", size=12,
               halign="center", valign="center", color="FFFFFF")
    ws.row_dimensions[1].height = 22

    # ── Baris 2: Instruksi ───────────────────────────────────────
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = (
        "Petunjuk: Baca Ground Truth 1 DAN Ground Truth 2 pada setiap baris. "
        "Beri skor di kolom KUNING untuk masing-masing Ground Truth.  "
        "1.0 = Sangat Lengkap & Akurat  |  0.8 = Akurat, kurang detail  |  "
        "0.5 = Sebagian benar  |  0.2 = Kurang akurat  |  0.0 = Tidak akurat / tidak relevan"
    )
    cell_style(ws["A2"], fill_hex="E8F5E9", size=12,
               halign="center", valign="center", color="1B5E20")
    ws.row_dimensions[2].height = 24

    # ── Baris 3: Header kolom ────────────────────────────────────
    header_dark = "263238"
    headers_main = [
        (COL_NO,  "No"),
        (COL_PYK, "Penyakit"),
        (COL_QRY, "Query / Pertanyaan"),
        (COL_GT1, "Ground Truth 1\n(dari Evaluasi RAG — acuan utama ROUGE/Cosine)"),
        (COL_GT2, "Ground Truth 2\n(dari Literatur Penyakit Padi — BB Padi/IRRI/Kementan)"),
    ]
    for col, label in headers_main:
        c = ws.cell(row=3, column=col, value=label)
        cell_style(c, bold=True, fill_hex=header_dark, size=12,
                   halign="center", valign="center", color="FFFFFF")

    # Header kolom skor GT (kuning, mencolok)
    skor_headers = [
        (COL_SKOR_GT1, "Skor\nGround Truth 1\n(ISI VALIDATOR)"),
        (COL_SKOR_GT2, "Skor\nGround Truth 2\n(ISI VALIDATOR)"),
    ]
    for col, label in skor_headers:
        c = ws.cell(row=3, column=col, value=label)
        cell_style(c, bold=True, fill_hex="F57F17", size=12,
                   halign="center", valign="center", color="FFFFFF")

    ws.row_dimensions[3].height = 50
    ws.freeze_panes = "A4"

    # ── Baris data ───────────────────────────────────────────────
    alt = ["FFFFFF", "F5F5F5"]

    for row_i, pq_ref in enumerate(first_gen, 1):
        row  = row_i + 3
        qid  = pq_ref["id"]
        dis  = pq_ref["disease"]
        fill = alt[row_i % 2]

        # GT-1
        gt1_raw  = pq_ref.get("ground_truth", "")
        gt1_body = gt1_raw if len(gt1_raw) <= MAX_CHARS_GT else \
                   gt1_raw[:MAX_CHARS_GT].rsplit(" ", 1)[0] + "…"
        kb_entry   = kb_data.get(dis.lower(), "")
        m_hdr      = re.match(r"\[(.+?)\]", kb_entry) if kb_entry else None
        gt1_header = m_hdr.group(1) if m_hdr else dis
        gt1 = f"[{gt1_header}]\n{gt1_body}"

        # GT-2
        gt2 = kb_data.get(dis.lower(), "—")

        data_cols = [
            (COL_NO,  row_i, "center"),
            (COL_PYK, dis,   "left"),
            (COL_QRY, pq_ref.get("query", ""), "left"),
            (COL_GT1, gt1,   "left"),
            (COL_GT2, gt2,   "left"),
        ]
        for col, val, ha in data_cols:
            c = ws.cell(row=row, column=col, value=val)
            cell_style(c, fill_hex=fill, size=12, halign=ha)

        # Kolom skor GT-1 dan GT-2 — kuning, diisi validator
        for col in [COL_SKOR_GT1, COL_SKOR_GT2]:
            c = ws.cell(row=row, column=col, value="")
            cell_style(c, fill_hex="FFF176", halign="center",
                       valign="center", size=12, bold=True)

        # Tinggi baris otomatis berdasarkan isi terpanjang
        def est_lines(text, col_width_chars):
            if not text:
                return 1
            chars_per_line = max(1, int(col_width_chars * 1.6))
            lines = str(text).splitlines()
            return sum(max(1, -(-len(ln) // chars_per_line)) for ln in lines)

        max_lines = max(
            est_lines(pq_ref.get("query", ""), col_widths[COL_QRY]),
            est_lines(gt1, col_widths[COL_GT1]),
            est_lines(gt2, col_widths[COL_GT2]),
        )
        ws.row_dimensions[row].height = max(60, min(max_lines * 14 + 8, 600))

    # ════════════════════════════════════════════════════════════
    # BAGIAN BAWAH: Catatan/Saran Tambahan + Tanda Tangan
    # ════════════════════════════════════════════════════════════
    last_data_row = len(first_gen) + 3

    # ── Spacer ────────────────────────────────────────────────────
    spacer_row = last_data_row + 1
    ws.row_dimensions[spacer_row].height = 8

    # ── Label catatan ─────────────────────────────────────────────
    note_label_row = spacer_row + 1
    ws.merge_cells(f"A{note_label_row}:{last_col}{note_label_row}")
    c_lbl = ws[f"A{note_label_row}"]
    c_lbl.value = "CATATAN / SARAN TAMBAHAN VALIDATOR:"
    cell_style(c_lbl, bold=True, fill_hex="263238", size=12,
               halign="left", valign="center", color="FFFFFF")
    ws.row_dimensions[note_label_row].height = 22

    # ── Area isi catatan (3 baris gabung, kosong untuk diisi) ─────
    note_start = note_label_row + 1
    note_end   = note_start + 2
    ws.merge_cells(f"A{note_start}:{last_col}{note_end}")
    c_note = ws[f"A{note_start}"]
    c_note.value = (
        "Tuliskan di sini jika ada kekurangan, informasi yang perlu ditambahkan, "
        "atau saran perbaikan terhadap Ground Truth yang ada:\n\n"
    )
    cell_style(c_note, fill_hex="FAFAFA", size=12, halign="left", valign="top")
    ws.row_dimensions[note_start].height = 90

    # ── Spacer sebelum TTD ────────────────────────────────────────
    ttd_spacer = note_end + 1
    ws.row_dimensions[ttd_spacer].height = 14

    # ── Baris 1: Kota & Tanggal (di atas tanda tangan, sisi kanan) ─
    ttd_kota_row = ttd_spacer + 1
    ws.merge_cells(f"E{ttd_kota_row}:{last_col}{ttd_kota_row}")
    c_kota = ws[f"E{ttd_kota_row}"]
    c_kota.value = "Indramayu, ……………………… 2026"
    c_kota.font      = Font(bold=False, size=12, name="Arial Narrow")
    c_kota.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[ttd_kota_row].height = 20

    # ── Baris 2: Label jabatan ────────────────────────────────────
    ttd_label_row = ttd_kota_row + 1
    ws.merge_cells(f"E{ttd_label_row}:{last_col}{ttd_label_row}")
    c_ttd_lbl = ws[f"E{ttd_label_row}"]
    c_ttd_lbl.value = "Validator / Petani,"
    c_ttd_lbl.font      = Font(bold=False, size=12, name="Arial Narrow")
    c_ttd_lbl.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[ttd_label_row].height = 20

    # ── Baris 3-5: Ruang tanda tangan ────────────────────────────
    for r in [ttd_label_row + 1, ttd_label_row + 2, ttd_label_row + 3]:
        ws.row_dimensions[r].height = 20

    # ── Baris 6: Garis tanda tangan ──────────────────────────────
    ttd_line_row = ttd_label_row + 3
    ws.merge_cells(f"E{ttd_line_row}:{last_col}{ttd_line_row}")
    c_line = ws[f"E{ttd_line_row}"]
    c_line.value = "( ………………………………………………… )"
    c_line.font      = Font(bold=False, size=12, name="Arial Narrow")
    c_line.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[ttd_line_row].height = 20

    # ── Baris 7: Nama & Jabatan ───────────────────────────────────
    ttd_name_row = ttd_line_row + 1
    ws.merge_cells(f"E{ttd_name_row}:{last_col}{ttd_name_row}")
    c_name = ws[f"E{ttd_name_row}"]
    c_name.value = "Nama & Jabatan Petani"
    c_name.font      = Font(bold=True, size=12, name="Arial Narrow", color="555555")
    c_name.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[ttd_name_row].height = 18

    total_rows = ttd_name_row
    ws.print_area = f"A1:{last_col}{total_rows}"

    # ════════════════════════════════════════════════════════════
    # SHEET 2: Panduan Validator
    # ════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Panduan untuk Validator")
    ws2.page_setup.orientation = "portrait"
    ws2.page_setup.paperSize   = 9
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 78

    ws2.merge_cells("A1:B1")
    ws2["A1"] = "PANDUAN PENGISIAN LEMBAR PENILAIAN FAITHFULNESS"
    cell_style(ws2["A1"], bold=True, fill_hex="1B5E20", size=12,
               halign="center", color="FFFFFF")
    ws2.row_dimensions[1].height = 26

    panduan = [
        ("", ""),
        ("APA YANG DINILAI?", ""),
        ("Faithfulness",
         "Seberapa lengkap dan akurat isi setiap Ground Truth dibandingkan "
         "dengan pengetahuan lapangan (gejala penyakit, nama obat/pestisida, "
         "dan cara penanganan yang benar menurut literatur BB Padi, IRRI, Kementan)."),
        ("", ""),
        ("2 SUMBER\nGROUND TRUTH", ""),
        ("Ground Truth 1",
         "Jawaban referensi yang dibuat peneliti berdasarkan literatur — "
         "digunakan sebagai acuan utama evaluasi otomatis (ROUGE, Cosine Similarity)."),
        ("Ground Truth 2",
         "Ringkasan GEJALA dan PENANGANAN dari file penyakit_padi.txt "
         "(BB Padi, IRRI, Kementan). Digunakan sebagai acuan tambahan."),
        ("", ""),
        ("CARA MENGISI", ""),
        ("Langkah 1", "Buka sheet 'Penilaian Faithfulness'"),
        ("Langkah 2",
         "Baca kolom 'Ground Truth 1' — perhatikan apakah gejala, nama penyakit, "
         "nama obat/pestisida, dan langkah penanganan sudah lengkap dan benar"),
        ("Langkah 3",
         "Baca kolom 'Ground Truth 2' — bandingkan isinya dengan GT-1 "
         "dan pengetahuan lapangan Bapak/Ibu"),
        ("Langkah 4",
         "Isi skor di kolom KUNING 'Skor GT-1' dan 'Skor GT-2' secara terpisah "
         "sesuai kualitas masing-masing"),
        ("Langkah 5",
         "Skor GT-1 dan GT-2 boleh berbeda — nilai sesuai kelengkapan "
         "dan ketepatan informasi masing-masing"),
        ("", ""),
        ("PANDUAN SKOR", ""),
        ("1.0",
         "SANGAT LENGKAP & AKURAT — semua fakta penting ada: nama penyakit, "
         "gejala khas, nama obat/pestisida spesifik (beserta dosis), "
         "dan langkah penanganan sesuai kondisi lapangan"),
        ("0.8",
         "AKURAT — sebagian besar benar, satu-dua detail kurang spesifik "
         "(misal: pestisida disebut tapi tanpa dosis) namun tidak menyesatkan"),
        ("0.5",
         "SEBAGIAN BENAR — ada informasi penting yang terlewat atau "
         "berbeda dari referensi (misal: gejala benar tapi penanganan kurang lengkap)"),
        ("0.2",
         "KURANG AKURAT — sebagian besar informasi berbeda, tidak lengkap, "
         "atau tidak relevan dengan kondisi lapangan"),
        ("0.0",
         "TIDAK AKURAT — isi tidak sesuai penyakit yang ditanya, "
         "atau informasi yang diberikan salah / menyesatkan petani"),
        ("", ""),
        ("CONTOH PENILAIAN", ""),
        ("Skor 1.0",
         "GT menyebut: 'Semprot Kocide 77 WP dosis 2-3 g/L, kurangi nitrogen, perbaiki drainase'\n"
         "→ Lengkap: ada nama pestisida, dosis, dan tindakan budidaya"),
        ("Skor 0.5",
         "GT menyebut: 'Jaga kebersihan lahan dan hindari kelembaban tinggi'\n"
         "→ Benar tapi tidak menyebut pestisida spesifik atau dosis"),
        ("Skor 0.0",
         "GT memberikan penjelasan penyakit lain (tidak relevan dengan query)\n"
         "→ Tidak sesuai sama sekali"),
        ("", ""),
        ("CATATAN",
         "• Tidak perlu mengisi seluruh tabel dalam satu waktu — bisa dicicil.\n"
         "• Baris 'healthy' dan 'harvest_stage' BUKAN penyakit — nilai berdasarkan "
         "kelengkapan panduan pemeliharaan/panen yang tersedia.\n"
         "• Jika ada GT yang menurut Bapak/Ibu perlu diperbaiki, beri catatan "
         "di sel skor (contoh: '0.5 — gejala kurang spesifik').\n"
         "• Setelah selesai, kirimkan kembali file Excel ini ke peneliti."),
    ]

    for i, (k, v) in enumerate(panduan, 2):
        ck = ws2.cell(row=i, column=1, value=k)
        cv = ws2.cell(row=i, column=2, value=v)

        if k in ("APA YANG DINILAI?", "CARA MENGISI", "PANDUAN SKOR",
                 "CONTOH PENILAIAN", "CATATAN", "2 SUMBER\nGROUND TRUTH"):
            cell_style(ck, bold=True, fill_hex="E8F5E9", color="1B5E20", size=12)
            cell_style(cv, bold=True, fill_hex="E8F5E9", color="1B5E20", size=12)
        elif k.startswith("Langkah"):
            cell_style(ck, bold=True, fill_hex="E3F2FD", size=12)
            cell_style(cv, fill_hex="E3F2FD", size=12)
        elif k in ("1.0", "0.8", "0.5", "0.2", "0.0"):
            cell_style(ck, bold=True, fill_hex="FFF9C4", halign="center", size=12)
            cell_style(cv, fill_hex="FFF9C4", size=12)
        elif k.startswith("Skor ") or k.startswith("Ground Truth") or k == "Faithfulness":
            cell_style(ck, bold=True, fill_hex="F5F5F5", size=12)
            cell_style(cv, fill_hex="F5F5F5", size=12)
        else:
            ws2.row_dimensions[i].height = 5
            continue

        ws2.row_dimensions[i].height = 55 if "\n" in str(v) else 20

    wb.save(output_path)

    print(f"\n✅ File Excel (v3 — Penilaian Ground Truth) berhasil dibuat: {output_path}")
    print(f"\nIsi file:")
    print(f"  Sheet 1 — 'Penilaian Faithfulness'")
    print(f"    {len(first_gen)} baris query")
    print(f"    Kolom: No | Penyakit | Query | GT-1 | GT-2 | Skor GT-1 | Skor GT-2")
    print(f"    GT-2 tersedia: {sum(1 for pq in first_gen if kb_data.get(pq['disease'].lower()))} dari {len(first_gen)} penyakit")
    print(f"  Sheet 2 — 'Panduan untuk Validator'")
    print(f"\nCara cetak: File → Print → A4 Landscape → Fit Sheet on One Page")


def main():
    parser = argparse.ArgumentParser(
        description="Buat Excel penilaian faithfulness (v3 — Penilaian Ground Truth)"
    )
    parser.add_argument("--input",  default=JSON_PATH, help="Path JSON hasil evaluasi")
    parser.add_argument("--kb",     default=KB_PATH,   help="Path knowledge base .txt")
    parser.add_argument("--output", default=OUTPUT,    help="Path output Excel")
    args = parser.parse_args()

    if not Path(args.input).exists():
        print(f"❌ File tidak ditemukan: {args.input}")
        print(f"   Jalankan dulu: python evaluate_rag.py --k 3 --llm all")
        return

    print(f"Membaca JSON    : {args.input}")
    d = load_json(args.input)

    kb_data = {}
    if Path(args.kb).exists():
        print(f"Membaca KB      : {args.kb}")
        kb_data = parse_kb(args.kb)
        print(f"GT-2 ditemukan  : {len(kb_data)} penyakit dari knowledge base")
    else:
        print(f"⚠️  KB tidak ditemukan ({args.kb}), GT-2 akan kosong")

    build_excel(d, kb_data, args.output)


if __name__ == "__main__":
    main()