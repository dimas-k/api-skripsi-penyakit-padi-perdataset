"""
hitung_metrik_setelah_validasi.py
===================================
Jalankan SETELAH petani/validator mengisi skor GT-1 dan GT-2 di Excel.

Alur:
  1. Baca Excel validator (kolom Skor GT-1 & Skor GT-2 per baris)
  2. Tentukan Ground Truth terbaik per query:
       - Jika Skor GT-1 > Skor GT-2  → pakai GT-1
       - Jika Skor GT-2 > Skor GT-1  → pakai GT-2
       - Jika sama (atau keduanya kosong) → gabungkan GT-1 + GT-2
  3. Hitung ROUGE-1 / ROUGE-2 / ROUGE-L / BLEU / METEOR
     antara GT terpilih vs jawaban LLM (LOW / MEDIUM / HIGH)
  4. Simpan hasil ke Excel (per query + ringkasan rata-rata per tier)

Cara pakai:
  python hitung_metrik_setelah_validasi.py
  python hitung_metrik_setelah_validasi.py --excel Penilaian_Faithfulness_untuk_Validator.xlsx
                                           --json   hasil_evaluasi_rag.json
                                           --output hasil_metrik_final.xlsx
"""

import json, re, argparse, math
from pathlib import Path
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Library metrik ─────────────────────────────────────────────────────────
from rouge_score import rouge_scorer
import nltk
from nltk.translate.bleu_score import sentence_bleu, SmoothingFunction
from nltk.translate.meteor_score import meteor_score
from nltk.tokenize import word_tokenize

nltk.download("punkt",     quiet=True)
nltk.download("punkt_tab", quiet=True)
nltk.download("wordnet",   quiet=True)

# ── Konfigurasi default ─────────────────────────────────────────────────────
EXCEL_PATH  = "Penilaian_Faithfulness_untuk_Validator.xlsx"
JSON_PATH   = "hasil_evaluasi_rag.json"
OUTPUT_PATH = "hasil_metrik_final.xlsx"

TIER_KEYWORDS = {
    "LOW"   : ["LOW",  "low",  "qwen", "Qwen",  "3b",  "3B"],
    "MEDIUM": ["MEDIUM", "medium", "Gemini", "gemini", "Flash"],
    "HIGH"  : ["HIGH", "high", "llama", "Llama", "70b", "70B",
               "versatile", "groq", "Groq"],
}

TIER_COLORS = {
    "LOW"   : ("1565C0", "BBDEFB"),   # (header, data)
    "MEDIUM": ("1A7A4A", "C8E6C9"),
    "HIGH"  : ("6A1B9A", "E1BEE7"),
}


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1 — BACA EXCEL VALIDATOR
# ══════════════════════════════════════════════════════════════════════════════
def read_validator_excel(excel_path: str) -> dict[str, dict]:
    """
    Baca sheet 'Penilaian Faithfulness'.
    Kembalikan dict keyed by ID query:
      {
        "Q01": {
          "id": "Q01",
          "disease": "bacterial_leaf_blight",
          "query": "...",
          "gt1": "...",
          "gt2": "...",
          "skor_gt1": 0.8,    # None jika belum diisi
          "skor_gt2": 0.5,
        },
        ...
      }
    """
    wb = load_workbook(excel_path, read_only=True)

    # Cari sheet yang benar
    sheet_name = None
    for name in wb.sheetnames:
        if "penilaian" in name.lower() or "faithfulness" in name.lower():
            sheet_name = name
            break
    if not sheet_name:
        raise ValueError(f"Sheet 'Penilaian Faithfulness' tidak ditemukan di {excel_path}")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    # Cari baris header (ada kata "No" di kolom pertama)
    header_row_idx = None
    for i, row in enumerate(rows):
        if row and str(row[0]).strip().lower() == "no":
            header_row_idx = i
            break

    if header_row_idx is None:
        raise ValueError("Baris header 'No' tidak ditemukan di Excel.")

    headers = [str(h).strip().lower() if h else "" for h in rows[header_row_idx]]

    # Temukan indeks kolom
    def find_col(keywords):
        for i, h in enumerate(headers):
            if any(kw.lower() in h for kw in keywords):
                return i
        return None

    col_no    = find_col(["no"])
    col_pyk   = find_col(["penyakit"])
    col_qry   = find_col(["query"])
    col_gt1   = find_col(["ground truth 1", "gt-1", "gt1", "dari evaluasi"])
    col_gt2   = find_col(["ground truth 2", "gt-2", "gt2", "literatur"])
    col_sgt1  = find_col(["skor\nground truth 1", "skor gt-1", "skor gt1", "skor\ngt1"])
    col_sgt2  = find_col(["skor\nground truth 2", "skor gt-2", "skor gt2", "skor\ngt2"])

    # Fallback: cari "skor" columns by position (kolom 6 dan 7 setelah header)
    if col_sgt1 is None:
        skor_cols = [i for i, h in enumerate(headers) if "skor" in h and i not in [col_no, col_pyk, col_qry, col_gt1, col_gt2]]
        if len(skor_cols) >= 2:
            col_sgt1, col_sgt2 = skor_cols[0], skor_cols[1]
        elif len(skor_cols) == 1:
            col_sgt1 = skor_cols[0]

    print(f"  Kolom ditemukan: No={col_no} Penyakit={col_pyk} Query={col_qry} "
          f"GT1={col_gt1} GT2={col_gt2} SkorGT1={col_sgt1} SkorGT2={col_sgt2}")

    result = {}
    for i, row in enumerate(rows[header_row_idx + 1:], 1):
        if not row or not row[col_no if col_no else 0]:
            continue

        def safe(col):
            if col is None or col >= len(row):
                return None
            return row[col]

        def to_float(val):
            try:
                return float(val) if val not in (None, "") else None
            except (TypeError, ValueError):
                return None

        entry = {
            "row_no"   : safe(col_no),
            "disease"  : str(safe(col_pyk) or "").strip(),
            "query"    : str(safe(col_qry) or "").strip(),
            "gt1"      : str(safe(col_gt1) or "").strip(),
            "gt2"      : str(safe(col_gt2) or "").strip(),
            "skor_gt1" : to_float(safe(col_sgt1)),
            "skor_gt2" : to_float(safe(col_sgt2)),
        }
        # ID query dari kolom No (Q01, Q02, ...) — skip baris non-angka (footer, catatan, dsb)
        row_no_val = str(safe(col_no) or "").strip()
        if not row_no_val.isdigit():
            continue
        qid = f"Q{int(row_no_val):02d}"
        result[qid] = entry

    wb.close()
    return result


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2 — PILIH GROUND TRUTH TERBAIK
# ══════════════════════════════════════════════════════════════════════════════
def pick_best_gt(entry: dict) -> tuple[str, str]:
    """
    Kembalikan (gt_text, gt_source) berdasarkan skor validator.
      gt_source: "GT-1" | "GT-2" | "GT-1+2 (digabung)"
    """
    s1 = entry.get("skor_gt1")
    s2 = entry.get("skor_gt2")
    gt1 = entry.get("gt1", "")
    gt2 = entry.get("gt2", "")

    # Jika belum diisi (None) → gabungkan
    if s1 is None and s2 is None:
        return (f"{gt1}\n{gt2}".strip(), "GT-1+2 (belum diisi)")

    if s1 is None:
        return (gt2, "GT-2")
    if s2 is None:
        return (gt1, "GT-1")

    if s1 > s2:
        return (gt1, f"GT-1 (skor {s1} > {s2})")
    elif s2 > s1:
        return (gt2, f"GT-2 (skor {s2} > {s1})")
    else:
        # Skor sama → gabungkan
        return (f"{gt1}\n{gt2}".strip(), f"GT-1+2 digabung (skor sama: {s1})")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 — HITUNG METRIK
# ══════════════════════════════════════════════════════════════════════════════
_rouge = rouge_scorer.RougeScorer(["rouge1", "rouge2", "rougeL"], use_stemmer=False)
_smooth = SmoothingFunction().method1

def clean_text(text: str) -> str:
    """Hapus markdown, normalisasi whitespace."""
    text = re.sub(r"\*\*|__|[*_]", "", text)
    text = re.sub(r"#+\s*", "", text)
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def hitung_rouge(reference: str, hypothesis: str) -> dict:
    scores = _rouge.score(reference, hypothesis)
    return {
        "rouge1_p": round(scores["rouge1"].precision, 4),
        "rouge1_r": round(scores["rouge1"].recall, 4),
        "rouge1_f": round(scores["rouge1"].fmeasure, 4),
        "rouge2_f": round(scores["rouge2"].fmeasure, 4),
        "rougeL_f": round(scores["rougeL"].fmeasure, 4),
    }


def hitung_bleu(reference: str, hypothesis: str) -> float:
    ref_tokens = word_tokenize(reference.lower())
    hyp_tokens = word_tokenize(hypothesis.lower())
    if not hyp_tokens or not ref_tokens:
        return 0.0
    score = sentence_bleu([ref_tokens], hyp_tokens, smoothing_function=_smooth)
    return round(score, 4)


def hitung_meteor(reference: str, hypothesis: str) -> float:
    ref_tokens = word_tokenize(reference.lower())
    hyp_tokens = word_tokenize(hypothesis.lower())
    if not hyp_tokens or not ref_tokens:
        return 0.0
    score = meteor_score([ref_tokens], hyp_tokens)
    return round(score, 4)


def compute_all_metrics(reference: str, hypothesis: str) -> dict:
    ref_clean = clean_text(reference)
    hyp_clean = clean_text(hypothesis)
    metrics = hitung_rouge(ref_clean, hyp_clean)
    metrics["bleu"]   = hitung_bleu(ref_clean, hyp_clean)
    metrics["meteor"] = hitung_meteor(ref_clean, hyp_clean)
    return metrics


# ══════════════════════════════════════════════════════════════════════════════
# STEP 4 — BACA JSON & GABUNGKAN
# ══════════════════════════════════════════════════════════════════════════════
def detect_tier(llm_name: str) -> str:
    for tier, kws in TIER_KEYWORDS.items():
        if any(kw in llm_name for kw in kws):
            return tier
    return "OTHER"


def load_json_data(json_path: str) -> dict:
    with open(json_path, encoding="utf-8") as f:
        return json.load(f)


def build_results(validator_data: dict, json_data: dict) -> list[dict]:
    """
    Gabungkan data validator + jawaban LLM, hitung metrik.
    Kembalikan list of dict (satu baris per query × tier).
    """
    gen_dict = {}

    # Support struktur: generators (list) atau generator (dict)
    if "generators" in json_data and isinstance(json_data["generators"], list):
        for g in json_data["generators"]:
            if "error" not in g:
                tier = detect_tier(g["llm"])
                gen_dict[tier] = {pq["id"]: pq for pq in g.get("per_query", [])}
    elif "generator" in json_data and isinstance(json_data["generator"], dict):
        for llm_key, g in json_data["generator"].items():
            if isinstance(g, dict) and "error" not in g:
                tier = detect_tier(g.get("llm", llm_key))
                gen_dict[tier] = {pq["id"]: pq for pq in g.get("per_query", [])}

    rows = []
    for qid, entry in sorted(validator_data.items()):
        gt_text, gt_source = pick_best_gt(entry)

        for tier in ["LOW", "MEDIUM", "HIGH"]:
            if tier not in gen_dict:
                continue
            pq = gen_dict[tier].get(qid)
            if not pq:
                continue

            generated = pq.get("generated", "")
            metrics   = compute_all_metrics(gt_text, generated)

            rows.append({
                "qid"        : qid,
                "disease"    : entry["disease"],
                "query"      : entry["query"],
                "tier"       : tier,
                "llm"        : pq.get("llm", tier),
                "gt_source"  : gt_source,
                "skor_gt1"   : entry.get("skor_gt1"),
                "skor_gt2"   : entry.get("skor_gt2"),
                "gt_used"    : gt_text,
                "generated"  : generated,
                **metrics,
            })

    return rows


# ══════════════════════════════════════════════════════════════════════════════
# STEP 5 — BUILD EXCEL OUTPUT
# ══════════════════════════════════════════════════════════════════════════════
def make_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)


def cs(cell, bold=False, fill_hex=None, size=10, halign="left",
       valign="center", color="000000", wrap=True):
    cell.font      = Font(bold=bold, size=size, name="Arial Narrow", color=color)
    cell.alignment = Alignment(horizontal=halign, vertical=valign,
                               wrap_text=wrap)
    cell.border    = make_border()
    if fill_hex:
        cell.fill  = PatternFill("solid", fgColor=fill_hex)


def build_excel_output(rows: list[dict], output_path: str):
    wb = Workbook()

    # ── SHEET 1: Detail per Query ─────────────────────────────────────────────
    ws = wb.active
    ws.title = "Detail per Query"

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize   = 9
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1

    COLS = [
        ("No",         5),
        ("ID",         6),
        ("Penyakit",   18),
        ("Query",      28),
        ("Tier",       9),
        ("GT Dipakai", 14),
        ("Skor GT-1",  10),
        ("Skor GT-2",  10),
        ("ROUGE-1 F",  10),
        ("ROUGE-2 F",  10),
        ("ROUGE-L F",  10),
        ("BLEU",       10),
        ("METEOR",     10),
    ]

    # Set lebar kolom
    for i, (_, w) in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Baris judul
    last_col = get_column_letter(len(COLS))
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "HASIL METRIK FAITHFULNESS — Setelah Validasi Ground Truth oleh Petani"
    cs(ws["A1"], bold=True, fill_hex="1B5E20", size=12,
       halign="center", color="FFFFFF", wrap=False)
    ws.row_dimensions[1].height = 22

    # Baris header
    for col_i, (label, _) in enumerate(COLS, 1):
        c = ws.cell(row=2, column=col_i, value=label)
        cs(c, bold=True, fill_hex="263238", size=9, halign="center", color="FFFFFF")
    ws.row_dimensions[2].height = 28
    ws.freeze_panes = "A3"

    # Data
    tier_fill_alt = {
        "LOW"   : ["BBDEFB", "E3F2FD"],
        "MEDIUM": ["C8E6C9", "E8F5E9"],
        "HIGH"  : ["E1BEE7", "F3E5F5"],
    }

    no = 0
    for r in rows:
        no += 1
        tier  = r["tier"]
        fills = tier_fill_alt.get(tier, ["FFFFFF", "F5F5F5"])
        fill  = fills[no % 2]

        vals = [
            no,
            r["qid"],
            r["disease"],
            r["query"],
            tier,
            r["gt_source"],
            r["skor_gt1"],
            r["skor_gt2"],
            r["rouge1_f"],
            r["rouge2_f"],
            r["rougeL_f"],
            r["bleu"],
            r["meteor"],
        ]
        aligns = ["center","center","left","left","center","left",
                  "center","center","center","center","center","center","center"]

        row_idx = no + 2
        for col_i, (val, ha) in enumerate(zip(vals, aligns), 1):
            c = ws.cell(row=row_idx, column=col_i, value=val)
            cs(c, fill_hex=fill, size=9, halign=ha)
        ws.row_dimensions[row_idx].height = 32

    # ── SHEET 2: Ringkasan Rata-rata per Tier ────────────────────────────────
    ws2 = wb.create_sheet("Ringkasan per Tier")

    ws2.merge_cells("A1:H1")
    ws2["A1"] = "RINGKASAN RATA-RATA METRIK PER TIER LLM"
    cs(ws2["A1"], bold=True, fill_hex="1B5E20", size=12,
       halign="center", color="FFFFFF", wrap=False)
    ws2.row_dimensions[1].height = 22

    sum_cols = [
        ("Tier",       12),
        ("Model LLM",  22),
        ("ROUGE-1 F",  12),
        ("ROUGE-2 F",  12),
        ("ROUGE-L F",  12),
        ("BLEU",       12),
        ("METEOR",     12),
        ("n Queries",  10),
    ]
    for i, (label, w) in enumerate(sum_cols, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
        c = ws2.cell(row=2, column=i, value=label)
        cs(c, bold=True, fill_hex="263238", size=10, halign="center", color="FFFFFF")
    ws2.row_dimensions[2].height = 26

    # Hitung rata-rata per tier
    from collections import defaultdict
    tier_agg = defaultdict(lambda: {
        "llm": "", "rouge1": [], "rouge2": [], "rougeL": [], "bleu": [], "meteor": []
    })
    for r in rows:
        t = r["tier"]
        tier_agg[t]["llm"]   = r["llm"]
        tier_agg[t]["rouge1"].append(r["rouge1_f"])
        tier_agg[t]["rouge2"].append(r["rouge2_f"])
        tier_agg[t]["rougeL"].append(r["rougeL_f"])
        tier_agg[t]["bleu"]  .append(r["bleu"])
        tier_agg[t]["meteor"].append(r["meteor"])

    for row_i, tier in enumerate(["LOW", "MEDIUM", "HIGH"], 3):
        if tier not in tier_agg:
            continue
        agg = tier_agg[tier]
        avg = lambda lst: round(sum(lst) / len(lst), 4) if lst else 0.0
        hdr_hex, data_hex = TIER_COLORS.get(tier, ("263238", "F5F5F5"))

        vals = [
            tier,
            agg["llm"],
            avg(agg["rouge1"]),
            avg(agg["rouge2"]),
            avg(agg["rougeL"]),
            avg(agg["bleu"]),
            avg(agg["meteor"]),
            len(agg["rouge1"]),
        ]
        aligns = ["center","left","center","center","center","center","center","center"]

        for col_i, (val, ha) in enumerate(zip(vals, aligns), 1):
            c = ws2.cell(row=row_i, column=col_i, value=val)
            cs(c, fill_hex=data_hex, size=10, halign=ha, bold=(col_i == 1))
        ws2.row_dimensions[row_i].height = 22

    # ── SHEET 3: GT yang Digunakan ──────────────────────────────────────────
    ws3 = wb.create_sheet("GT yang Digunakan")
    ws3.merge_cells("A1:F1")
    ws3["A1"] = "GROUND TRUTH YANG DIPILIH BERDASARKAN VALIDASI PETANI"
    cs(ws3["A1"], bold=True, fill_hex="1B5E20", size=12,
       halign="center", color="FFFFFF", wrap=False)
    ws3.row_dimensions[1].height = 22

    gt_cols = [
        ("ID",        6),
        ("Penyakit",  18),
        ("Skor GT-1", 10),
        ("Skor GT-2", 10),
        ("GT Dipilih",14),
        ("Teks GT yang Dipakai", 70),
    ]
    for i, (label, w) in enumerate(gt_cols, 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
        c = ws3.cell(row=2, column=i, value=label)
        cs(c, bold=True, fill_hex="263238", size=9, halign="center", color="FFFFFF")
    ws3.row_dimensions[2].height = 26

    seen_qids = set()
    row_gt = 3
    for r in rows:
        if r["qid"] in seen_qids:
            continue
        seen_qids.add(r["qid"])
        fill = "F5F5F5" if row_gt % 2 == 0 else "FFFFFF"
        gt_vals = [
            r["qid"], r["disease"], r["skor_gt1"], r["skor_gt2"],
            r["gt_source"], r["gt_used"][:300] + "…" if len(r["gt_used"]) > 300 else r["gt_used"]
        ]
        aligns = ["center","left","center","center","left","left"]
        for col_i, (val, ha) in enumerate(zip(gt_vals, aligns), 1):
            c = ws3.cell(row=row_gt, column=col_i, value=val)
            cs(c, fill_hex=fill, size=9, halign=ha)
        ws3.row_dimensions[row_gt].height = 55
        row_gt += 1

    wb.save(output_path)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(
        description="Hitung ROUGE/BLEU/METEOR setelah validasi GT oleh petani"
    )
    parser.add_argument("--excel",  default=EXCEL_PATH,  help="Excel yang sudah diisi validator")
    parser.add_argument("--json",   default=JSON_PATH,   help="JSON hasil evaluasi RAG")
    parser.add_argument("--output", default=OUTPUT_PATH, help="Path output Excel hasil metrik")
    args = parser.parse_args()

    # Validasi file
    for path, label in [(args.excel, "Excel validator"), (args.json, "JSON evaluasi")]:
        if not Path(path).exists():
            print(f"❌ {label} tidak ditemukan: {path}")
            return

    print(f"\n{'='*60}")
    print("  HITUNG METRIK SETELAH VALIDASI GROUND TRUTH")
    print(f"{'='*60}")

    # Step 1: Baca Excel
    print(f"\n[1/4] Membaca Excel validator: {args.excel}")
    validator_data = read_validator_excel(args.excel)
    print(f"      {len(validator_data)} query ditemukan")

    # Cek status pengisian skor
    filled = sum(1 for e in validator_data.values()
                 if e.get("skor_gt1") is not None or e.get("skor_gt2") is not None)
    print(f"      {filled}/{len(validator_data)} query sudah ada skor validator")
    if filled == 0:
        print("  ⚠️  Skor validator belum diisi — GT terbaik akan ditentukan otomatis (gabungan GT-1+GT-2)")

    # Step 2: Pilih GT terbaik (diprint untuk cek)
    print(f"\n[2/4] Menentukan Ground Truth terbaik per query:")
    for qid, entry in sorted(validator_data.items()):
        _, source = pick_best_gt(entry)
        s1 = entry.get("skor_gt1", "—")
        s2 = entry.get("skor_gt2", "—")
        print(f"      {qid} [{entry['disease'][:20]:20s}] → {source}  (s1={s1}, s2={s2})")

    # Step 3: Baca JSON + hitung metrik
    print(f"\n[3/4] Membaca JSON & menghitung metrik: {args.json}")
    json_data = load_json_data(args.json)
    rows = build_results(validator_data, json_data)
    print(f"      {len(rows)} baris metrik dihitung ({len(validator_data)} query × 3 tier)")

    # Tampilkan ringkasan
    print(f"\n[4/4] Ringkasan rata-rata per Tier:")
    from collections import defaultdict
    tier_agg = defaultdict(lambda: {"r1": [], "r2": [], "rL": [], "bleu": [], "meteor": []})
    for r in rows:
        t = r["tier"]
        tier_agg[t]["r1"]    .append(r["rouge1_f"])
        tier_agg[t]["r2"]    .append(r["rouge2_f"])
        tier_agg[t]["rL"]    .append(r["rougeL_f"])
        tier_agg[t]["bleu"]  .append(r["bleu"])
        tier_agg[t]["meteor"].append(r["meteor"])

    avg = lambda lst: sum(lst) / len(lst) if lst else 0.0
    print(f"\n  {'Tier':<8} {'ROUGE-1':>8} {'ROUGE-2':>8} {'ROUGE-L':>8} {'BLEU':>8} {'METEOR':>8}")
    print(f"  {'-'*52}")
    for tier in ["LOW", "MEDIUM", "HIGH"]:
        if tier not in tier_agg:
            continue
        a = tier_agg[tier]
        print(f"  {tier:<8} "
              f"{avg(a['r1']):>8.4f} "
              f"{avg(a['r2']):>8.4f} "
              f"{avg(a['rL']):>8.4f} "
              f"{avg(a['bleu']):>8.4f} "
              f"{avg(a['meteor']):>8.4f}")

    # Simpan Excel
    build_excel_output(rows, args.output)
    print(f"\n✅ Hasil disimpan ke: {args.output}")
    print(f"   Sheet 1: Detail per Query (semua metrik per baris)")
    print(f"   Sheet 2: Ringkasan Rata-rata per Tier")
    print(f"   Sheet 3: GT yang Digunakan (untuk laporan)")


if __name__ == "__main__":
    main()