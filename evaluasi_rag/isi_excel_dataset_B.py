"""
isi_excel_dataset_B.py  (Dataset B — Validator Dosen Pertanian Unwir)
=====================================================================
Membangun lembar penilaian untuk validator pakar ke-2 (dosen pertanian Unwir),
berbeda dari Dataset A (Q01-Q18 yang sudah divalidasi Pak Tangono / WIL).

Struktur kolom (8 kolom, A4 Landscape):
  A=No | B=Penyakit | C=Query | D=GT-1 | E=GT-2 |
  F=Skor GT-1 | G=Skor GT-2 | H=Validasi Pakar (KOLOM KOSONG -> JADI GT-3)

18 pertanyaan baru (Q19-Q36) mengambil ANGLE berbeda dari Dataset A:
  - Pencegahan, varietas tahan, pengelolaan air, dosis pupuk,
    timing aplikasi, alternatif organik, diagnosis banding, dll.

GT-1 ditulis ulang lebih lengkap (target validator beri skor 1):
  5 elemen: penyebab, gejala, penanganan kuratif, pencegahan, varietas tahan.

GT-2 diekstrak dari knowledge_base/penyakit_padi.txt dengan parser yang lebih
lengkap (gabung GEJALA + PENANGANAN + PENCEGAHAN + KONDISI PENYEBARAN).

Cara pakai:
  python isi_excel_dataset_B.py
  python isi_excel_dataset_B.py --kb knowledge_base/penyakit_padi.txt --out Penilaian_Faithfulness_DatasetB.xlsx
"""

import re, argparse
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.datavalidation import DataValidation


# ═════════════════════════════════════════════════════════════════
# 18 PERTANYAAN BARU (Q19-Q36) + GT-1 LENGKAP (developer-written)
# Angle BEDA dari Dataset A (Q01-Q18 yang divalidasi Pak Tangono).
# ═════════════════════════════════════════════════════════════════
from dataset_b_questions import DATASET_B


# ═════════════════════════════════════════════════════════════════
# PARSE KB → GT-2 (lebih lengkap: GEJALA + PENANGANAN + PENCEGAHAN + KONDISI)
# ═════════════════════════════════════════════════════════════════
def extract_section(blok: str, section_pattern: str, max_len: int = 400) -> str:
    m = re.search(section_pattern, blok, re.S | re.I)
    if not m:
        return ""
    text = re.sub(r"\s+", " ", m.group(1).strip())
    if len(text) > max_len:
        text = text[:max_len].rsplit(" ", 1)[0] + "..."
    return text


def parse_kb_full(kb_path: str) -> dict[str, str]:
    """GT-2 lebih lengkap: gejala, penanganan, pencegahan, kondisi penyebaran."""
    text = Path(kb_path).read_text(encoding="utf-8")
    bloks = re.split(r"={3,}", text)

    result: dict[str, str] = {}
    current_kode = None
    current_nama = None

    for blok in bloks:
        m_kode = re.search(r"Kode:\s*(\w+)", blok)
        if m_kode:
            current_kode = m_kode.group(1).strip().lower()

        m_nama = re.search(r"PENYAKIT:\s*(.+?)(?:\(|\n)", blok)
        if m_nama:
            current_nama = m_nama.group(1).strip()

        if not current_kode:
            continue

        gejala     = extract_section(blok, r"GEJALA\s*:\s*(.*?)(?=\n[A-Z]{3,}|\Z)")
        kondisi    = extract_section(blok, r"KONDISI[^:]*:\s*(.*?)(?=\n[A-Z]{3,}|\Z)", max_len=300)
        penanganan_raw = extract_section(
            blok,
            r"PENANG[A-Z]+(?:\s+[A-Z]+)*\s*:\s*(.*?)(?=\n[A-Z]{4,}|\Z)",
            max_len=500,
        )
        pencegahan = extract_section(blok, r"PENCEGAHAN\s*:\s*(.*?)(?=\n[A-Z]{3,}|\Z)", max_len=300)

        if not (gejala or penanganan_raw or pencegahan):
            continue

        header = f"{current_nama} ({current_kode})" if current_nama else current_kode
        parts = [f"[{header}]"]
        if gejala:
            parts.append(f"Gejala: {gejala}")
        if kondisi:
            parts.append(f"Kondisi Penyebaran: {kondisi}")
        if penanganan_raw:
            parts.append(f"Penanganan: {penanganan_raw}")
        if pencegahan:
            parts.append(f"Pencegahan: {pencegahan}")

        result[current_kode] = "\n".join(parts).strip()

        current_kode = None
        current_nama = None

    return result


# ═════════════════════════════════════════════════════════════════
# STYLING
# ═════════════════════════════════════════════════════════════════
def make_border(style_type="thin"):
    s = Side(style=style_type)
    return Border(left=s, right=s, top=s, bottom=s)


def cell_style(cell, bold=False, fill_hex=None, size=12, wrap=True,
               halign="left", valign="top", color="000000"):
    cell.font = Font(bold=bold, size=size, name="Arial Narrow", color=color)
    cell.alignment = Alignment(horizontal=halign, vertical=valign,
                               wrap_text=wrap, shrink_to_fit=False)
    cell.border = make_border()
    if fill_hex:
        cell.fill = PatternFill("solid", fgColor=fill_hex)


# ═════════════════════════════════════════════════════════════════
# BUILD EXCEL DATASET B (8 KOLOM)
# ═════════════════════════════════════════════════════════════════
def build_excel(kb_data: dict, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Penilaian Faithfulness B"

    # A4 Landscape
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = 9
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.scale = 100
    ws.print_options.horizontalCentered = True
    ws.page_margins = PageMargins(
        left=0.4, right=0.4, top=0.5, bottom=0.5,
        header=0.2, footer=0.2,
    )
    ws.oddHeader.center.text = "LEMBAR PENILAIAN FAITHFULNESS (Dataset B) — Sistem RAG Penyakit Padi"
    ws.oddFooter.right.text = "Halaman &P dari &N"

    # ── 8 kolom ──
    COL_NO        = 1
    COL_PYK       = 2
    COL_QRY       = 3
    COL_GT1       = 4
    COL_GT2       = 5
    COL_SKOR_GT1  = 6
    COL_SKOR_GT2  = 7
    COL_VAL_PAKAR = 8   # KOLOM KOSONG → akan jadi GT-3
    TOTAL_COLS    = 8

    col_widths = {
        COL_NO       : 4,
        COL_PYK      : 18,
        COL_QRY      : 28,
        COL_GT1      : 45,
        COL_GT2      : 45,
        COL_SKOR_GT1 : 11,
        COL_SKOR_GT2 : 11,
        COL_VAL_PAKAR: 55,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    last_col = get_column_letter(TOTAL_COLS)

    # ── Baris 1: Judul ──
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = "LEMBAR PENILAIAN FAITHFULNESS (Dataset B) — Sistem Rekomendasi Penyakit Padi"
    cell_style(ws["A1"], bold=True, fill_hex="1B5E20", size=12,
               halign="center", valign="center", color="FFFFFF")
    ws.row_dimensions[1].height = 22

    # ── Baris 2: Instruksi ──
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = (
        "Petunjuk: (1) Baca Ground Truth 1 dan Ground Truth 2 pada setiap baris. "
        "(2) Beri skor di kolom KUNING untuk masing-masing GT (1.0=Sangat Lengkap & Akurat, "
        "0.8=Akurat kurang detail, 0.5=Sebagian benar, 0.2=Kurang akurat, 0.0=Tidak akurat). "
        "(3) Tuliskan VALIDASI PAKAR Bapak/Ibu di kolom BIRU (jawaban referensi versi pakar) — "
        "jawaban ini akan menjadi Ground Truth ke-3."
    )
    cell_style(ws["A2"], fill_hex="E8F5E9", size=11,
               halign="center", valign="center", color="1B5E20")
    ws.row_dimensions[2].height = 38

    # ── Baris 3: Header kolom ──
    header_dark = "263238"
    headers_main = [
        (COL_NO,  "No"),
        (COL_PYK, "Penyakit"),
        (COL_QRY, "Query / Pertanyaan"),
        (COL_GT1, "Ground Truth 1\n(Disusun Peneliti — acuan literatur)"),
        (COL_GT2, "Ground Truth 2\n(Dari Knowledge Base — BB Padi/IRRI/Kementan)"),
    ]
    for col, label in headers_main:
        c = ws.cell(row=3, column=col, value=label)
        cell_style(c, bold=True, fill_hex=header_dark, size=11,
                   halign="center", valign="center", color="FFFFFF")

    skor_headers = [
        (COL_SKOR_GT1, "Skor\nGT-1\n(ISI PAKAR)"),
        (COL_SKOR_GT2, "Skor\nGT-2\n(ISI PAKAR)"),
    ]
    for col, label in skor_headers:
        c = ws.cell(row=3, column=col, value=label)
        cell_style(c, bold=True, fill_hex="F57F17", size=11,
                   halign="center", valign="center", color="FFFFFF")

    # Header kolom validasi pakar (biru, mencolok) → akan jadi GT-3
    c = ws.cell(
        row=3, column=COL_VAL_PAKAR,
        value="VALIDASI PAKAR\n(Jawaban Referensi versi Pakar)\n→ Ground Truth 3"
    )
    cell_style(c, bold=True, fill_hex="0D47A1", size=11,
               halign="center", valign="center", color="FFFFFF")

    ws.row_dimensions[3].height = 60
    ws.freeze_panes = "A4"

    # ── Isi data Q19-Q36 ──
    start_row = 4
    for i, item in enumerate(DATASET_B):
        r = start_row + i
        kode = item["disease"]
        gt2 = kb_data.get(kode, "(GT-2 tidak tersedia di KB)")

        cells_data = [
            (COL_NO,  i + 1,         "center", "center", False, None),
            (COL_PYK, kode,          "left",   "top",    False, None),
            (COL_QRY, item["query"], "left",   "top",    False, None),
            (COL_GT1, item["gt1"],   "left",   "top",    False, None),
            (COL_GT2, gt2,           "left",   "top",    False, None),
            (COL_SKOR_GT1, "",       "center", "center", False, "FFF9C4"),  # kuning lembut
            (COL_SKOR_GT2, "",       "center", "center", False, "FFF9C4"),
            (COL_VAL_PAKAR, "",      "left",   "top",    False, "E3F2FD"),  # biru lembut
        ]
        for col, val, ha, va, bold, fill in cells_data:
            c = ws.cell(row=r, column=col, value=val)
            cell_style(c, bold=bold, fill_hex=fill, size=11,
                       halign=ha, valign=va)

        ws.row_dimensions[r].height = 180

    # ── Data validation: skor 0.0-1.0 ──
    last_data_row = start_row + len(DATASET_B) - 1
    dv = DataValidation(
        type="decimal",
        operator="between",
        formula1=0.0,
        formula2=1.0,
        showErrorMessage=True,
        errorTitle="Skor tidak valid",
        error="Isi nilai antara 0.0 sampai 1.0",
    )
    dv.add(f"{get_column_letter(COL_SKOR_GT1)}{start_row}:{get_column_letter(COL_SKOR_GT2)}{last_data_row}")
    ws.add_data_validation(dv)

    # ── Baris terakhir: TTD pakar ──
    sig_row = last_data_row + 2
    ws.merge_cells(f"A{sig_row}:H{sig_row}")
    ws[f"A{sig_row}"] = "PERNYATAAN VALIDATOR PAKAR"
    cell_style(ws[f"A{sig_row}"], bold=True, fill_hex="263238", size=12,
               halign="center", valign="center", color="FFFFFF")
    ws.row_dimensions[sig_row].height = 22

    sig_lines = [
        "Nama Lengkap & Gelar : ___________________________________________________",
        "NIDN / Jabatan        : ___________________________________________________",
        "Bidang Keahlian       : ___________________________________________________",
        "Instansi              : Universitas Wiralodra Indramayu (Unwir)",
        "Tanggal Penilaian     : ___________________________________________________",
        "",
        "Tanda Tangan          :",
        "",
        "                        ___________________________________________________",
    ]
    for j, line in enumerate(sig_lines):
        rr = sig_row + 1 + j
        ws.merge_cells(f"A{rr}:H{rr}")
        ws[f"A{rr}"] = line
        cell_style(ws[f"A{rr}"], size=11, halign="left", valign="center")
        ws.row_dimensions[rr].height = 22

    # ════════════════════════════════════════════════════════════
    # SHEET 2: Panduan untuk Pakar
    # ════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Panduan untuk Pakar")
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 90

    ws2.merge_cells("A1:B1")
    ws2["A1"] = "PANDUAN PENILAIAN — Validator Pakar (Dosen Pertanian Unwir)"
    cell_style(ws2["A1"], bold=True, fill_hex="1B5E20", size=14,
               halign="center", valign="center", color="FFFFFF")
    ws2.row_dimensions[1].height = 28

    panduan = [
        ("Konteks Penelitian",
         "Skripsi mahasiswa D4 Rekayasa Perangkat Lunak Politeknik Negeri Indramayu — "
         "membangun aplikasi mobile PadiCare yang memberi rekomendasi penanganan penyakit padi "
         "otomatis menggunakan AI (deep learning + Large Language Model)."),
        ("Tujuan Validasi",
         "Menilai keakuratan dua jawaban referensi (Ground Truth) untuk 18 pertanyaan teknis "
         "penyakit padi, dan menyediakan jawaban versi pakar sebagai referensi tambahan."),
        ("Tugas Pakar",
         "(1) Membaca pasangan pertanyaan + GT-1 + GT-2 pada setiap baris. "
         "(2) Memberi skor 0.0 - 1.0 untuk masing-masing GT pada kolom KUNING. "
         "(3) Menuliskan jawaban referensi versi pakar di kolom BIRU (Validasi Pakar). "
         "Jawaban pakar ini akan dijadikan Ground Truth ke-3."),
        ("Rubrik Skor",
         "1.0 = Sangat lengkap & sangat akurat (tidak ada kekurangan)\n"
         "0.8 = Akurat namun kurang detail (informasi inti benar, ada poin yang bisa diperdalam)\n"
         "0.5 = Sebagian benar (ada poin yang kurang tepat atau hilang signifikan)\n"
         "0.2 = Kurang akurat (banyak kesalahan atau informasi tidak lengkap)\n"
         "0.0 = Tidak akurat / tidak relevan dengan pertanyaan"),
        ("Tips Menulis Validasi Pakar",
         "Jawaban tidak harus panjang. Mohon mencakup elemen berikut bila relevan: "
         "penyebab/patogen, gejala kunci, penanganan kuratif (bahan aktif & dosis), "
         "pencegahan, dan varietas tahan bila ada. Boleh tulis tangan setelah dicetak "
         "atau ketik di file digital."),
        ("Kerahasiaan & Pencantuman Nama",
         "Nama, gelar, dan kredensial Bapak/Ibu akan dicantumkan di lampiran skripsi "
         "sebagai narasumber pakar / validator independen. Tidak ada data pribadi lain "
         "yang akan dipublikasikan."),
        ("Estimasi Waktu",
         "Sekitar 1-2 jam pertemuan, atau bisa dititipkan dan dijemput kembali setelah "
         "Bapak/Ibu sempat menyelesaikan."),
        ("Kontak Mahasiswa",
         "Dimas Arya Ramadhan Setiawan — NIM 2205067\n"
         "D4 Rekayasa Perangkat Lunak — Politeknik Negeri Indramayu\n"
         "Email: 2205067@student.polindra.ac.id"),
    ]
    for i, (k, v) in enumerate(panduan):
        r = 2 + i
        ws2[f"A{r}"] = k
        ws2[f"B{r}"] = v
        cell_style(ws2[f"A{r}"], bold=True, fill_hex="E8F5E9", size=11,
                   halign="left", valign="top", color="1B5E20")
        cell_style(ws2[f"B{r}"], size=11, halign="left", valign="top")
        ws2.row_dimensions[r].height = 70

    wb.save(output_path)
    print(f"✓ Disimpan: {output_path}")
    print(f"  Total pertanyaan: {len(DATASET_B)}")
    print(f"  Kolom: 8 (No, Penyakit, Query, GT-1, GT-2, Skor GT-1, Skor GT-2, Validasi Pakar/GT-3)")


# ═════════════════════════════════════════════════════════════════
# MAIN
# ═════════════════════════════════════════════════════════════════
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--kb", default="knowledge_base/penyakit_padi.txt")
    ap.add_argument("--out", default="hasil/Penilaian_Faithfulness_DatasetB.xlsx")
    args = ap.parse_args()

    kb_data = parse_kb_full(args.kb)
    print(f"✓ KB diparse: {len(kb_data)} kode penyakit")

    build_excel(kb_data, args.out)


if __name__ == "__main__":
    main()
