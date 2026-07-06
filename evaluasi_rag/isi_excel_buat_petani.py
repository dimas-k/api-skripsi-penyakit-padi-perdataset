import json, re, argparse
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

# ─── Konfigurasi ──────────────────────────────────
JSON_PATH  = "hasil/hasil_evaluasi_rag.json"
KB_PATH    = "knowledge_base/penyakit_padi.txt"
OUTPUT     = "hasil/Penilaian_Faithfulness_Pakar_petani.xlsx"
DATASET_LABEL = "B"          # label dataset yang muncul di judul & nama sheet

MAX_CHARS_GT = 1000  # max karakter ground truth per kolom

INSTANSI_PAKAR = "Universitas Wiralodra Indramayu (Unwir)"
KONTAK_MAHASISWA = (
    "Dimas Arya Ramadhan Setiawan — NIM 2205067\n"
    "D4 Rekayasa Perangkat Lunak — Politeknik Negeri Indramayu\n"
    "Email: 2205067@student.polindra.ac.id"
)

TIER_KEYWORDS = {
    "LOW"   : ["LOW",  "low",  "qwen", "Qwen", "3b", "3B"],
    "MEDIUM": ["MEDIUM", "medium", "Gemini", "gemini", "Flash"],
    "HIGH"  : ["HIGH", "high", "llama", "Llama", "70b", "70B",
            "versatile", "groq", "Groq"],
}


# ════════════════════════════════════════════════════
# PARSE KNOWLEDGE BASE → Ground Truth ke-2
# ════════════════════════════════════════════════════
def parse_kb(kb_path: str) -> dict:
    """Baca penyakit_padi.txt dan ekstrak GEJALA + PENANGANAN per kode penyakit."""
    text = Path(kb_path).read_text(encoding="utf-8")
    bloks = re.split(r"={3,}", text)

    result = {}
    kode    = None
    nama_id = None

    for blok in bloks:
        m = re.search(r"Kode:\s*(\w+)", blok)
        if m:
            kode = m.group(1).strip().lower()

        mn = re.search(r"PENYAKIT\s*:\s*(.+)", blok, re.I)
        if mn:
            # ambil nama Indonesia sebelum tanda kurung, mis. 'HAWAR DAUN BAKTERI'
            nama_id = re.sub(r"\s*\(.*?\)\s*:?\s*$", "", mn.group(1).strip()).strip().rstrip(":").strip()

        if not kode:
            continue

        # Ekstrak satu seksi (dari beberapa kandidat nama header) dan
        # rapikan spasi. Berhenti di header seksi berikutnya (baris KAPITAL
        # diakhiri titik dua) atau akhir blok.
        def _section(*names):
            for nm in names:
                mm = re.search(
                    rf"{nm}\s*:\s*(.*?)(?=\n\s*(?:DESKRIPSI|GEJALA|CIRI-CIRI|INDIKATOR|KONDISI PENYEBARAN|PENYEBARAN|PENANGA[HN]AN DAN PENGENDALIAN|PENANGA[HN]AN|PENGENDALIAN|PEMELIHARAAN|PENCEGAHAN|TINGKAT KEPARAHAN|KAITAN|CATATAN|REFERENSI)[^\n]*:|\Z)",
                    blok, re.S,
                )
                if mm and mm.group(1).strip():
                    return re.sub(r"\s+", " ", mm.group(1)).strip()
            return ""

        gejala     = _section("GEJALA", "CIRI-CIRI", "INDIKATOR")
        kondisi    = _section("KONDISI PENYEBARAN", "PENYEBARAN")
        penanganan = _section("PENANGA[HN]AN DAN PENGENDALIAN", "PENANGA[HN]AN",
                              "PENGENDALIAN", "PEMELIHARAAN")
        pencegahan = _section("PENCEGAHAN")

        if gejala or kondisi or penanganan or pencegahan:
            header = f"{nama_id} ({kode})" if nama_id else kode
            parts = [f"[{header}]"]
            if gejala:
                parts.append(f"Gejala: {gejala}")
            if kondisi:
                parts.append(f"Kondisi Penyebaran: {kondisi}")
            if penanganan:
                parts.append(f"Penanganan: {penanganan}")
            if pencegahan:
                parts.append(f"Pencegahan: {pencegahan}")
            result[kode] = "\n".join(parts).strip()
            kode    = None
            nama_id = None

    return result


# ════════════════════════════════════════════════════
# UTILS
# ════════════════════════════════════════════════════
def load_json(path: str) -> dict:
    return json.loads(Path(path).read_text(encoding="utf-8"))


def extract_generators(d: dict) -> list:
    raw = []
    if "generators" in d and isinstance(d["generators"], list):
        raw = [g for g in d["generators"] if "error" not in g]
    elif "generator" in d and isinstance(d["generator"], dict):
        raw = [g for g in d["generator"].values()
               if isinstance(g, dict) and "error" not in g]

    def tier_order(g):
        name = g.get("llm", "")
        for i, (tier, kws) in enumerate(TIER_KEYWORDS.items()):
            if any(kw in name for kw in kws):
                return i
        return 99

    return sorted(raw, key=tier_order)


# ════════════════════════════════════════════════════
# STYLING
# ════════════════════════════════════════════════════
def make_border(style_type="thin"):
    s = Side(style=style_type)
    return Border(left=s, right=s, top=s, bottom=s)


def cell_style(cell, bold=False, fill_hex=None, size=12, wrap=True,
               halign="left", valign="top", color="000000"):
    cell.font      = Font(bold=bold, size=size, name="Arial Narrow", color=color)
    cell.alignment = Alignment(horizontal=halign, vertical=valign,
                               wrap_text=wrap, shrink_to_fit=False)
    cell.border    = make_border()
    if fill_hex:
        cell.fill = PatternFill("solid", fgColor=fill_hex)


# ════════════════════════════════════════════════════
# BUILD EXCEL
# ════════════════════════════════════════════════════
def build_excel(d: dict, kb_data: dict, output_path: str, dataset_label: str = DATASET_LABEL):
    generators = extract_generators(d)
    if not generators:
        print("❌ Tidak ada data generator di JSON.")
        return

    first_gen = generators[0]["per_query"]

    wb = Workbook()

    # ═══════════════════════════════════════════
    # SHEET 1: Penilaian Faithfulness (Dataset X)
    # ═══════════════════════════════════════════
    ws = wb.active
    ws.title = f"Penilaian Faithfulness {dataset_label}".strip()

    ws.page_setup.orientation    = "landscape"
    ws.page_setup.paperSize      = 9
    ws.page_setup.fitToPage      = True
    ws.page_setup.fitToWidth     = 1
    ws.page_setup.fitToHeight    = 0
    ws.page_setup.scale          = 100
    ws.print_options.horizontalCentered = True
    ws.page_margins = PageMargins(
        left=0.4, right=0.4, top=0.5, bottom=0.5,
        header=0.2, footer=0.2
    )
    ws.oddHeader.center.text = "LEMBAR PENILAIAN FAITHFULNESS — Sistem RAG Penyakit Padi"
    ws.oddFooter.right.text  = "Halaman &P dari &N"

    # ── Struktur kolom (8 kolom) ───────────────────────
    COL_NO       = 1
    COL_PYK      = 2
    COL_QRY      = 3
    COL_GT1      = 4
    COL_GT2      = 5
    COL_SKOR_GT1 = 6
    COL_SKOR_GT2 = 7
    COL_VALIDASI = 8   # GT-3: jawaban referensi versi pakar (BIRU, diisi pakar)
    TOTAL_COLS   = 8

    col_widths = {
        COL_NO       : 4,
        COL_PYK      : 18,
        COL_QRY      : 28,
        COL_GT1      : 45,
        COL_GT2      : 55,
        COL_SKOR_GT1 : 11,
        COL_SKOR_GT2 : 11,
        COL_VALIDASI : 55,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w

    last_col = get_column_letter(TOTAL_COLS)

    # ── Baris 1: Judul ──────────────────────────────
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = (f"LEMBAR PENILAIAN FAITHFULNESS (Dataset {dataset_label}) "
                f"— Sistem Rekomendasi Penyakit Padi")
    cell_style(ws["A1"], bold=True, fill_hex="1B5E20", size=12,
               halign="center", valign="center", color="FFFFFF")
    ws.row_dimensions[1].height = 22

    # ── Baris 2: Instruksi ───────────────────────────
    ws.merge_cells(f"A2:{last_col}2")
    ws["A2"] = (
        "Petunjuk: (1) Baca Ground Truth 1 dan Ground Truth 2 pada setiap baris. "
        "(2) Beri skor di kolom KUNING untuk masing-masing GT "
        "(1.0=Sangat Lengkap & Akurat, 0.8=Akurat kurang detail, 0.5=Sebagian benar, "
        "0.2=Kurang akurat, 0.0=Tidak akurat). "
        "(3) Tuliskan VALIDASI PAKAR Bapak/Ibu di kolom BIRU "
        "(jawaban referensi versi pakar) — jawaban ini akan menjadi Ground Truth ke-3."
    )
    cell_style(ws["A2"], fill_hex="E8F5E9", size=11,
               halign="center", valign="center", color="1B5E20")
    ws.row_dimensions[2].height = 40

    # ── Baris 3: Header kolom ────────────────────────
    header_dark = "263238"
    headers_main = [
        (COL_NO,  "No"),
        (COL_PYK, "Penyakit"),
        (COL_QRY, "Query / Pertanyaan"),
        (COL_GT1, "Ground Truth 1\n(Disusun Peneliti — acuan literatur)"),
        (COL_GT2, "Ground Truth 2\n(Dari Knowledge Base — BB Padi/IRRI/Kementan)"),
    ]
    for col, label in headers_main:
        c = ws.cell(row=3, column=col, value=label)
        cell_style(c, bold=True, fill_hex=header_dark, size=11,
                   halign="center", valign="center", color="FFFFFF")

    # Header kolom skor GT (kuning)
    skor_headers = [
        (COL_SKOR_GT1, "Skor\nGT-1\n(ISI PAKAR)"),
        (COL_SKOR_GT2, "Skor\nGT-2\n(ISI PAKAR)"),
    ]
    for col, label in skor_headers:
        c = ws.cell(row=3, column=col, value=label)
        cell_style(c, bold=True, fill_hex="F57F17", size=11,
                   halign="center", valign="center", color="FFFFFF")

    # Header kolom Validasi Pakar / GT-3 (biru)
    c_val = ws.cell(row=3, column=COL_VALIDASI,
                    value="VALIDASI PAKAR\n(Jawaban Referensi versi Pakar)\n→ Ground Truth 3")
    cell_style(c_val, bold=True, fill_hex="1565C0", size=11,
               halign="center", valign="center", color="FFFFFF")

    ws.row_dimensions[3].height = 52
    ws.freeze_panes = "A4"

    # ── Baris data ───────────────────────────────
    alt = ["FFFFFF", "F5F5F5"]

    for row_i, pq_ref in enumerate(first_gen, 1):
        row  = row_i + 3
        dis  = pq_ref["disease"]
        fill = alt[row_i % 2]

        gt1_raw  = pq_ref.get("ground_truth", "")
        gt1_body = gt1_raw if len(gt1_raw) <= MAX_CHARS_GT else \
                   gt1_raw[:MAX_CHARS_GT].rsplit(" ", 1)[0] + "…"
        # GT-1 ditampilkan apa adanya (tanpa prefix [kode]) agar identik
        # dengan Ground Truth 1 yang disusun peneliti di GROUND_TRUTH_QA
        gt1 = gt1_body

        gt2 = kb_data.get(dis.lower(), "—")

        data_cols = [
            (COL_NO,  row_i, "center"),
            (COL_PYK, dis,   "left"),
            (COL_QRY, pq_ref.get("query", ""), "left"),
            (COL_GT1, gt1,   "left"),
            (COL_GT2, gt2,   "left"),
        ]
        for col, val, ha in data_cols:
            c = ws.cell(row=row, column=col, value=val)
            cell_style(c, fill_hex=fill, size=11, halign=ha)

        # Kolom skor GT-1 & GT-2 — kuning, diisi pakar
        for col in [COL_SKOR_GT1, COL_SKOR_GT2]:
            c = ws.cell(row=row, column=col, value="")
            cell_style(c, fill_hex="FFF176", halign="center",
                       valign="center", size=12, bold=True)

        # Kolom Validasi Pakar (GT-3) — biru muda, diisi pakar
        cvp = ws.cell(row=row, column=COL_VALIDASI, value="")
        cell_style(cvp, fill_hex="E3F2FD", halign="left",
                   valign="top", size=11)

        def est_lines(text, col_width_chars):
            if not text:
                return 1
            chars_per_line = max(1, int(col_width_chars * 1.6))
            lines = str(text).splitlines()
            return sum(max(1, -(-len(ln) // chars_per_line)) for ln in lines)

        max_lines = max(
            est_lines(pq_ref.get("query", ""), col_widths[COL_QRY]),
            est_lines(gt1, col_widths[COL_GT1]),
            est_lines(gt2, col_widths[COL_GT2]),
        )
        ws.row_dimensions[row].height = max(60, min(max_lines * 14 + 8, 600))

    # ═══════════════════════════════════════════
    # BAGIAN BAWAH: PERNYATAAN VALIDATOR PAKAR
    # ═══════════════════════════════════════════
    last_data_row = len(first_gen) + 3
    r = last_data_row + 1
    ws.row_dimensions[r].height = 8   # spacer

    r += 1
    ws.merge_cells(f"A{r}:{last_col}{r}")
    c = ws[f"A{r}"]
    c.value = "PERNYATAAN VALIDATOR PAKAR"
    cell_style(c, bold=True, fill_hex="263238", size=12,
               halign="left", valign="center", color="FFFFFF")
    ws.row_dimensions[r].height = 22

    pernyataan = [
        "Nama Lengkap & Gelar : ___________________________________________________",
        "NIDN / Jabatan        : ___________________________________________________",
        "Bidang Keahlian       : ___________________________________________________",
        f"Instansi              : {INSTANSI_PAKAR}",
        "Tanggal Penilaian     : ___________________________________________________",
    ]
    for line in pernyataan:
        r += 1
        ws.merge_cells(f"A{r}:{last_col}{r}")
        c = ws[f"A{r}"]
        c.value = line
        c.font      = Font(bold=False, size=12, name="Arial Narrow")
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        ws.row_dimensions[r].height = 20

    r += 1
    ws.row_dimensions[r].height = 10  # spacer

    r += 1
    ws.merge_cells(f"A{r}:{last_col}{r}")
    c = ws[f"A{r}"]
    c.value = "Tanda Tangan          :"
    c.font      = Font(bold=False, size=12, name="Arial Narrow")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 20

    r += 1
    ws.row_dimensions[r].height = 26  # ruang TTD

    r += 1
    ws.merge_cells(f"A{r}:{last_col}{r}")
    c = ws[f"A{r}"]
    c.value = "                        ___________________________________________________"
    c.font      = Font(bold=False, size=12, name="Arial Narrow")
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[r].height = 20

    ws.print_area = f"A1:{last_col}{r}"

    # ═══════════════════════════════════════════
    # SHEET 2: Panduan untuk Pakar
    # ═══════════════════════════════════════════
    ws2 = wb.create_sheet("Panduan untuk Pakar")
    ws2.page_setup.orientation = "portrait"
    ws2.page_setup.paperSize   = 9
    ws2.column_dimensions["A"].width = 28
    ws2.column_dimensions["B"].width = 90

    ws2.merge_cells("A1:B1")
    ws2["A1"] = "PANDUAN PENILAIAN — Validator Pakar (Dosen Pertanian Unwir)"
    cell_style(ws2["A1"], bold=True, fill_hex="1B5E20", size=12,
               halign="center", color="FFFFFF")
    ws2.row_dimensions[1].height = 26

    panduan = [
        ("Konteks Penelitian",
         "Skripsi mahasiswa D4 Rekayasa Perangkat Lunak Politeknik Negeri Indramayu — "
         "membangun aplikasi mobile PadiCare yang memberi rekomendasi penanganan penyakit "
         "padi otomatis menggunakan AI (deep learning + Large Language Model)."),
        ("Tujuan Validasi",
         "Menilai keakuratan dua jawaban referensi (Ground Truth) untuk 18 pertanyaan teknis "
         "penyakit padi, dan menyediakan jawaban versi pakar sebagai referensi tambahan."),
        ("Tugas Pakar",
         "(1) Membaca pasangan pertanyaan + GT-1 + GT-2 pada setiap baris. "
         "(2) Memberi skor 0.0 - 1.0 untuk masing-masing GT pada kolom KUNING. "
         "(3) Menuliskan jawaban referensi versi pakar di kolom BIRU (Validasi Pakar). "
         "Jawaban pakar ini akan dijadikan Ground Truth ke-3."),
        ("Rubrik Skor",
         "1.0 = Sangat lengkap & sangat akurat (tidak ada kekurangan)\n"
         "0.8 = Akurat namun kurang detail (informasi inti benar, ada poin yang bisa diperdalam)\n"
         "0.5 = Sebagian benar (ada poin yang kurang tepat atau hilang signifikan)\n"
         "0.2 = Kurang akurat (banyak kesalahan atau informasi tidak lengkap)\n"
         "0.0 = Tidak akurat / tidak relevan dengan pertanyaan"),
        ("Tips Menulis Validasi Pakar",
         "Jawaban tidak harus panjang. Mohon mencakup elemen berikut bila relevan: "
         "penyebab/patogen, gejala kunci, penanganan kuratif (bahan aktif & dosis), "
         "pencegahan, dan varietas tahan bila ada. Boleh tulis tangan setelah dicetak "
         "atau ketik di file digital."),
        ("Kerahasiaan & Pencantuman Nama",
         "Nama, gelar, dan kredensial Bapak/Ibu akan dicantumkan di lampiran skripsi sebagai "
         "narasumber pakar / validator independen. Tidak ada data pribadi lain yang akan "
         "dipublikasikan."),
        ("Estimasi Waktu",
         "Sekitar 1-2 jam pertemuan, atau bisa dititipkan dan dijemput kembali setelah "
         "Bapak/Ibu sempat menyelesaikan."),
        ("Kontak Mahasiswa", KONTAK_MAHASISWA),
    ]

    row_i = 2
    for k, v in panduan:
        ck = ws2.cell(row=row_i, column=1, value=k)
        cv = ws2.cell(row=row_i, column=2, value=v)
        cell_style(ck, bold=True, fill_hex="E8F5E9", color="1B5E20", size=12,
                   halign="left", valign="top")
        cell_style(cv, fill_hex="F5F5F5", size=12, halign="left", valign="top")
        n_lines = str(v).count("\n") + max(1, len(str(v)) // 90)
        ws2.row_dimensions[row_i].height = max(30, n_lines * 15 + 8)
        row_i += 1

    wb.save(output_path)

    n_gt2 = sum(1 for pq in first_gen if kb_data.get(pq["disease"].lower()))
    print(f"\n✅ File Excel (v4 — Penilaian oleh PAKAR) berhasil dibuat: {output_path}")
    print(f"\nIsi file:")
    print(f"  Sheet 1 — 'Penilaian Faithfulness {dataset_label}'")
    print(f"    {len(first_gen)} baris query")
    print(f"    Kolom: No | Penyakit | Query | GT-1 | GT-2 | Skor GT-1 | Skor GT-2 | Validasi Pakar (GT-3)")
    print(f"    GT-2 tersedia: {n_gt2} dari {len(first_gen)} penyakit")
    print(f"  Sheet 2 — 'Panduan untuk Pakar'")
    print(f"\nCara cetak: File → Print → A4 Landscape → Fit Sheet on One Page")


# ═══ GROUND TRUTH 2 (disematkan persis dari Penilaian_Faithfulness_Petani.xlsx) ═══
# Sumber valid: nilai GT-2 di bawah identik dengan Excel petani agar tidak ada selisih.
GROUND_TRUTH_2 = {
    'bacterial_leaf_blight': '[HAWAR DAUN BAKTERI (bacterial_leaf_blight)]\nGejala: Gejala awal berupa bercak basah dan berminyak pada tepi daun yang kemudian meluas menjadi hawar (blight) berwarna kuning pucat hingga putih keabu-abuan. Gejala berkembang dari ujung dan tepi daun ke arah pangkal. Pada kondisi lembab, eksudat bakteri berwarna kekuningan (bacterial ooze) terlihat pada permukaan daun di pagi hari. Daun yang terinfeksi parah mengering dan berwarna seperti jerami. Serangan berat dapat menyebabkan seluruh daun mati.\nKondisi Penyebaran: Penyakit ini berkembang pesat pada kondisi suhu udara 25-35°C, kelembaban udara tinggi >85%, intensitas hujan tinggi, angin kencang, pemupukan nitrogen berlebihan, dan persawahan dengan genangan air. Kondisi sensor IoT yang mendukung penyebaran: suhu udara 28-32°C, kelembaban udara >80%, curah hujan >10 mm/hari.\nPenanganan: 1. Pengendalian Kimiawi: Semprot dengan bakterisida berbahan aktif tembaga (copper hydroxide) seperti Kocide 77 WP dosis 2-3 g/L air, atau Agrept 20 WP (streptomycin sulfate) dosis 1-2 g/L air. Lakukan penyemprotan setiap 7-10 hari sekali saat gejala mulai muncul. 2. Pengendalian Hayati: Aplikasikan Bacillus subtilis atau Pseudomonas fluorescens sebagai agen hayati dengan dosis 10 g/L air. 3. Pengendalian Budidaya: Kurangi dosis pupuk nitrogen, perbaiki drainase sawah, tanam varietas tahan HDB seperti IR64, Ciherang, atau Mekongga. 4. Sanitasi: Cabut dan musnahkan tanaman yang terinfeksi berat untuk mencegah penyebaran. 5. Pengaturan air: Hindari penggenangan berlebih, atur irigasi berselang (intermittent irrigation).\nPencegahan: Gunakan benih sehat bersertifikat, rendam benih dalam larutan bakterisida sebelum tanam, tanam varietas tahan, atur jarak tanam optimal 25x25 cm atau jajar legowo, hindari pemupukan nitrogen berlebihan, dan jaga kebersihan lahan dari sisa tanaman terinfeksi.',
    'leaf_blast': '[BLAS DAUN (leaf_blast)]\nGejala: Lesi berbentuk belah ketupat atau jarum dengan ujung runcing pada kedua sisinya (diamond-shaped lesion). Pusat lesi berwarna abu-abu atau putih dengan tepi berwarna coklat gelap dan zona kuning di sekitarnya (halo). Pada serangan berat, lesi menyatu sehingga daun tampak seperti terbakar. Bercak awal berwarna hijau abu-abu kecil yang berkembang menjadi lesi nekrotik berukuran 0.5-1.5 cm. Infeksi berat menyebabkan daun mengering dari pucuk ke bawah.\nKondisi Penyebaran: Lesi berbentuk belah ketupat atau jarum dengan ujung runcing pada kedua sisinya (diamond-shaped lesion). Pusat lesi berwarna abu-abu atau putih dengan tepi berwarna coklat gelap dan zona kuning di sekitarnya (halo). Pada serangan berat, lesi menyatu sehingga daun tampak seperti terbakar. Bercak awal berwarna hijau abu-abu kecil yang berkembang menjadi lesi nekrotik berukuran 0.5-1.5 cm. Infeksi berat menyebabkan daun mengering dari pucuk ke bawah.\nPenanganan: 1. Pengendalian Kimiawi (Prioritas Utama): Semprot dengan fungisida tricyclazole (Beam 75 WP dosis 0.5-1 g/L air), atau isoprothiolane (Fuji-one 400 EC dosis 1.5-2 ml/L), atau kasugamycin+tembaga (Blascide dosis 2 ml/L). Penyemprotan dilakukan segera saat gejala pertama muncul dan diulang 7-10 hari kemudian. 2. Pengendalian Hayati: Aplikasikan Trichoderma asperellum atau Bacillus subtilis sebagai bioagen preventif. 3. Pemupukan: Kurangi dosis nitrogen, tambahkan pupuk kalium (K) dan silika untuk meningkatkan ketahanan tanaman. 4. Pengaturan Irigasi: Kurangi penggenangan pada malam hari untuk mengurangi kelembaban kanopi. 5. Sanitasi: Musnahkan tanaman sakit dan sisa jerami untuk mengurangi sumber inokulum.\nPencegahan: Tanam varietas tahan blas (Ciherang Blast, Inpari 13, Inpari 19), hindari pemupukan nitrogen berlebih, semprot fungisida preventif pada saat primordia daun, jaga drainase yang baik, dan hindari menanam pada musim dengan kelembaban sangat tinggi tanpa perlindungan fungisida.',
    'neck_blast': '[BLAS LEHER MALAI (neck_blast)]\nGejala: Lesi berwarna coklat gelap hingga hitam pada leher malai (node pertama di bawah malai) atau pada buku-buku batang. Jaringan leher malai membusuk sehingga malai patah atau malai menggantung (patah leher). Gabah pada malai yang terinfeksi menjadi hampa, berwarna coklat atau putih kering. Malai tampak berdiri tegak meski tidak berisi (whitehead) atau menggantung lemah ke bawah. Beda dengan dead heart: dead heart terjadi pada fase vegetatif, neck blast terjadi pada fase generatif (pembungaan-pengisian biji).\nKondisi Penyebaran: Kondisi paling kritis: infeksi saat malai keluar (50% heading) dengan kelembaban >90%, suhu 24-28°C, angin, dan pemupukan nitrogen berlebih menjelang pembungaan. Ini adalah periode paling rentan tanaman.\nPenanganan: 1. TINDAKAN DARURAT - Semprot Segera: Gunakan fungisida tricyclazole (Beam 75 WP dosis 1-1.5 g/L) atau isoprothiolane (Fuji-one 400 EC dosis 2 ml/L) SEGERA saat gejala muncul. Penyemprotan tepat waktu sangat kritis - keterlambatan 3-4 hari dapat menyebabkan kehilangan hasil besar. 2. Pencegahan pada Primordia: Semprot fungisida preventif 10-14 hari sebelum malai keluar (fase primordia) dan pada saat 50% malai keluar (heading 50%). 3. Pemupukan: HENTIKAN pemberian nitrogen saat menjelang pembungaan. Tambahkan kalium dan silika untuk menguatkan dinding sel. 4. Pengairan: Pertahankan penggenangan dangkal untuk menjaga suhu dan kelembaban yang sesuai.\nPencegahan: Semprot fungisida preventif 2 kali: pertama saat primordia (7-10 HST sebelum malai keluar) dan kedua saat 50% heading. Gunakan varietas tahan blas. Hindari nitrogen berlebih menjelang pembungaan.',
    'sheath_blight': '[BUSUK PELEPAH (sheath_blight)]\nGejala: Lesi oval hingga tidak beraturan berwarna hijau abu-abu hingga putih keabu-abuan dengan tepi coklat pada pelepah daun (sheath) dekat permukaan air. Lesi berkembang dari pelepah bawah ke atas menuju daun. Pada kondisi lembab, miselium putih jamur terlihat pada permukaan lesi. Sklerotia berwarna coklat berbentuk tidak beraturan (seperti butiran tanah kecil) menempel pada permukaan lesi dan dapat jatuh ke air/tanah. Serangan berat menyebabkan seluruh pelepah dan daun membusuk, malai hampa atau tidak terbentuk sempurna.\nKondisi Penyebaran: Berkembang optimal pada suhu 28-32°C, kelembaban >85%, penggenangan air, dan tanam terlalu rapat. Pemupukan nitrogen berlebih sangat mendorong perkembangan penyakit. Kondisi sensor IoT yang mendukung: suhu 28-32°C, kelembaban udara >80%, nitrogen tanah tinggi.\nPenanganan: 1. Pengendalian Kimiawi: Semprot fungisida validamycin (Validacin 3L dosis 1-2 ml/L), hexaconazole (Anvil 50 SC dosis 1 ml/L), atau propiconazole (Tilt 250 EC dosis 0.5 ml/L) langsung ke pangkal tanaman (target pelepah). Penyemprotan volume tinggi (high volume spraying) lebih efektif agar larutan mencapai pangkal batang. 2. Pengelolaan Air: Kurangi penggenangan, gunakan irigasi berselang untuk memutus siklus infeksi. 3. Pemupukan: Kurangi dosis nitrogen, tingkatkan kalium untuk meningkatkan ketahanan. 4. Pengaturan Jarak Tanam: Perlebar jarak tanam (30x30 cm atau jajar legowo 2:1) untuk meningkatkan sirkulasi udara dan mengurangi kelembaban mikro. 5. Sanitasi: Kumpulkan dan musnahkan sklerotia setelah panen, bajak jerami yang terinfeksi.\nPencegahan: Hindari tanam terlalu rapat, kurangi dosis nitrogen, perbaiki drainase, gunakan varietas dengan ketahanan moderat, dan aplikasikan fungisida preventif saat kondisi mendukung penyakit (suhu tinggi + kelembaban tinggi).',
    'tungro': '[TUNGRO (tungro)]\nGejala: Daun berwarna kuning hingga kuning-oranye dimulai dari ujung daun, berbeda dengan defisiensi hara yang menguning dari pangkal daun. Tanaman kerdil (stunting) dengan tinggi hanya 1/3 hingga 1/2 dari tanaman normal. Jumlah anakan berkurang drastis. Daun muda menjadi sempit dan lebih pendek dari normal. Malai tidak terbentuk atau sangat sedikit dan hampa. Gejala mirip defisiensi nitrogen tetapi lebih parah dan tidak merespons pemupukan.\nKondisi Penyebaran: Penyebaran sangat tergantung pada populasi wereng hijau (vektor). Populasi wereng meningkat pada: awal musim hujan, pertanaman yang tidak serempak, pemupukan nitrogen berlebih, dan lahan dengan bayah (ratoon) atau gulma inang. Suhu 20-30°C dan kelembaban sedang mendukung perkembangan wereng.\nPenanganan: 1. PRIORITAS UTAMA - Kendalikan Vektor (Wereng Hijau): Semprot insektisida untuk membunuh wereng: imidakloprid (Confidor 200 SL dosis 0.5 ml/L), buprofezin (Applaud 25 WP dosis 1 g/L), atau BPMC (Bassa 500 EC dosis 2 ml/L). Pengendalian wereng harus dilakukan SEBELUM virus menyebar ke tanaman sehat. 2. Pencabutan Tanaman Sakit: Cabut dan musnahkan tanaman yang terinfeksi parah untuk mengurangi sumber virus (virus source). Kubur atau bakar tanaman yang dicabut. 3. Penanaman Serempak: Tanam serempak dalam satu kawasan untuk memutus ketersediaan inang bagi wereng sepanjang tahun. 4. Tidak Ada Pengobatan Langsung: Tidak ada fungisida atau bakterisida untuk penyakit virus. Fokus utama adalah MENGENDALIKAN WERENG dan MENCABUT TANAMAN SAKIT. 5. Tanam Ulang: Jika serangan >50% rumpun, pertimbangkan tanam ulang dengan varietas tahan tungro.\nPencegahan: Gunakan varietas tahan tungro (Tukad Unda, Tukad Petanu, Bondoyudo, Inpari 7), tanam serempak, kendalikan wereng hijau secara rutin dengan insektisida atau musuh alami (laba-laba, kepik predator), hindari tanam padi sepanjang tahun tanpa jeda (break crop), dan cabut tanaman inang (gulma padi liar) di sekitar sawah.',
    'brown_spot': '[BERCAK COKLAT (brown_spot)]\nGejala: Gejala berupa bercak oval hingga bundar berwarna coklat dengan ukuran 0,5-1 cm pada daun, pelepah, dan gabah. Bercak memiliki pusat berwarna abu-abu atau putih yang dikelilingi cincin berwarna coklat tua hingga ungu kecoklatan. Pada serangan berat, bercak-bercak menyatu sehingga daun tampak terbakar (blight). Bercak juga dapat muncul pada gabah menyebabkan biji berubah warna dan kualitas beras menurun.\nKondisi Penyebaran: Berkembang pesat pada lahan defisiensi kalium (K) dan silika, suhu 16-36°C (optimum 25-30°C), kelembaban >80%, dan lahan dengan kesuburan rendah. Kondisi sensor: kelembaban tinggi, pH tanah rendah, nitrogen dan kalium rendah.\nPenanganan: 1. Pengendalian Kimiawi: Semprot dengan fungisida berbahan aktif mancozeb (Dithane M-45 dosis 2 g/L), tricyclazole (Beam 75 WP dosis 0.5-1 g/L), atau propiconazole (Tilt 250 EC dosis 0.5 ml/L). Lakukan penyemprotan 2-3 kali dengan interval 10-14 hari. 2. Pemupukan: Tambahkan pupuk kalium (KCl) dosis 50-75 kg/ha dan pupuk silikat. Perbaiki keseimbangan hara dengan penambahan pupuk organik. 3. Pengendalian Hayati: Aplikasikan Trichoderma harzianum sebagai agen hayati. 4. Sanitasi: Bakar atau kubur sisa tanaman sakit untuk mengurangi sumber inokulum.\nPencegahan: Gunakan varietas tahan, lakukan pemupukan berimbang (N:P:K sesuai kebutuhan), tambahkan pupuk organik untuk memperbaiki struktur tanah, gunakan benih sehat, dan lakukan pergiliran tanaman.',
    'dead_heart': '[BATANG MATI / SUNDEP (dead_heart)]\nGejala: Gejala khas adalah pucuk/tunas yang mati berwarna kuning kecoklatan yang mudah dicabut karena batang bagian dalam telah dimakan larva. Daun tengah tanaman muda layu dan mati sementara daun luar masih hijau. Pada serangan parah, sebagian besar rumpun mengalami kematian pucuk (dead heart) sehingga tampak botak/gundul. Sering terlihat lubang kecil bekas gerekan pada batang dan kotoran larva berbentuk serbuk coklat.\nKondisi Penyebaran: Populasi penggerek meningkat pada kelembaban tinggi, periode tanpa hujan panjang diselingi hujan ringan, dan pada tanaman yang dipupuk nitrogen berlebihan. Suhu optimum untuk perkembangan larva: 25-30°C.\nPenanganan: 1. Pengendalian Insektisida: Gunakan insektisida karbofuran (Furadan 3G) dengan cara tabur 17 kg/ha pada saat 7-10 hari setelah tanam. Atau semprot dengan klorpirifos (Dursban 20 EC dosis 2 ml/L), fipronil (Regent 50 SC dosis 0.5-1 ml/L), atau imidakloprid (Confidor 200 SL dosis 0.5 ml/L) pada saat imago (ngengat) terlihat. 2. Pengendalian Hayati: Lepaskan parasitoid telur Trichogramma japonicum dengan dosis 200.000 parasitoid/ha/minggu. Gunakan Bacillus thuringiensis (Bt) sebagai insektisida hayati. 3. Pemasangan Lampu Perangkap: Pasang light trap untuk menangkap imago penggerek batang pada malam hari, terutama saat periode penerbangan (flight period). 4. Sanitasi: Cabut dan musnahkan tanaman yang terserang dead heart, bakar jerami setelah panen. 5. Pengaturan Waktu Tanam: Tanam serempak dalam satu hamparan untuk memutus siklus hama.\nPencegahan: Tanam serempak dan serentak dengan petani sekitar, gunakan varietas tahan penggerek, hindari pemupukan nitrogen berlebih, semprot insektisida saat telur ditemukan (sebelum menetas), dan pertahankan musuh alami seperti laba-laba dan kumbang.',
    'hispa': '[HISPA PADI (hispa)]\nGejala: Kerusakan imago: goresan horizontal berwarna putih sejajar (linear scratches) pada permukaan daun yang memberikan penampilan bergaris-garis putih. Goresan terjadi karena imago memakan klorofil daun dari permukaan atas. Kerusakan larva: blotch atau terowongan berwarna putih transparan pada helai daun karena larva menggerek jaringan daun dari dalam mesofil. Area yang digerek tampak seperti jendela transparan (windowing effect). Serangan berat menyebabkan daun mengering berwarna putih.\nKondisi Penyebaran: Populasi meningkat pada sawah dengan tanaman rapat, pemupukan nitrogen berlebih, dan musim hujan dengan kelembaban tinggi. Migrasi dari lahan rumput ke sawah terjadi pada awal musim tanam.\nPenanganan: 1. Pengendalian Insektisida: Semprot dengan insektisida klorpirifos (Dursban 20 EC dosis 2 ml/L), imidakloprid (Confidor 200 SL dosis 0.5 ml/L), atau deltametrin (Decis 25 EC dosis 0.5-1 ml/L). Penyemprotan pagi hari lebih efektif. 2. Pengendalian Mekanis: Celupkan daun yang terinfeksi ke dalam air atau usap dengan kain basah untuk menghilangkan imago dan telur. Potong daun yang menunjukkan tanda penggerekan larva. 3. Pengendalian Hayati: Pertahankan musuh alami seperti laba-laba, kumbang predator, dan parasitoid. 4. Pengaturan Jarak Tanam: Kurangi kepadatan tanaman untuk mengurangi kelembaban mikro dan ketersediaan inang.\nPencegahan: Tanam serempak dalam hamparan luas, hindari pemupukan nitrogen berlebih, pertahankan dan tingkatkan populasi musuh alami, dan lakukan pemantauan dini pada awal musim tanam terutama saat tanaman berumur 1-4 minggu.',
    'bacterial_panicle_blight': '[HAWAR MALAI BAKTERI (bacterial_panicle_blight)]\nGejala: Gejala utama berupa malai yang berdiri tegak karena gabah hampa (tidak berisi). Biji padi yang terinfeksi berwarna coklat hingga abu-abu dan mengeriput. Infeksi terjadi saat pembungaan, menyebabkan gabah gagal terisi. Pada serangan berat, seluruh malai dapat terlihat hampa dan berwarna coklat. Pangkal gabah berwarna coklat dengan batas yang jelas antara bagian sehat dan sakit.\nKondisi Penyebaran: Berkembang optimal pada suhu tinggi 30-35°C terutama saat pembungaan, kelembaban tinggi, dan kondisi angin. Infeksi terjadi pada saat anthesis (pembungaan) sehingga serangan pada periode ini sangat kritis.\nPenanganan: 1. Pengendalian Kimiawi: Semprot dengan bakterisida berbahan aktif kasugamycin seperti Kasumin 2L dosis 1-2 ml/L air pada saat primordial malai dan awal pembungaan. Atau gunakan oxolinic acid (Staimer 200 EC). 2. Pengendalian Budidaya: Hindari pemupukan nitrogen berlebih, atur waktu tanam agar pembungaan tidak bertepatan dengan suhu terlalu tinggi (>35°C). 3. Perlakuan benih: Rendam benih dalam air panas 52-54°C selama 10 menit atau dalam larutan bakterisida untuk meminimalkan infeksi terbawa benih. 4. Pengelolaan air: Irigasi malam hari untuk menurunkan suhu kanopi saat pembungaan.\nPencegahan: Gunakan benih sehat bersertifikat, lakukan perlakuan benih sebelum semai, pilih waktu tanam yang tepat agar pembungaan berlangsung saat suhu tidak terlalu tinggi, dan hindari pemupukan nitrogen berlebihan menjelang pembungaan.',
    'leaf_smut': '[GOSONG PALSU DAUN (leaf_smut)]\nGejala: Bercak-bercak kecil berwarna hitam atau abu-abu gelap berbentuk oval hingga memanjang tersebar pada helai daun. Bercak berukuran 1-5 mm dengan serbuk hitam (spora) di permukaan. Daun yang terinfeksi berat menunjukkan banyak titik-titik hitam yang menyebar di seluruh permukaan daun. Bercak tidak memiliki halo kuning yang jelas seperti blas, tetapi tepi bercak terkadang berwarna kuning pucat. Infeksi berat dapat menyebabkan daun menguning secara tidak merata.\nKondisi Penyebaran: Berkembang pada kondisi lembab dan suhu sedang 20-28°C. Spora tersebar melalui angin dan air.\nPenanganan: 1. Pengendalian Kimiawi: Semprot dengan fungisida tembaga (copper hydroxide) seperti Kocide 77 WP dosis 2 g/L, mancozeb (Dithane M-45 dosis 2 g/L), atau propiconazole (Tilt 250 EC dosis 0.5 ml/L). 2. Pengendalian Budidaya: Perbaiki drainase, hindari kepadatan tanaman berlebih, dan jaga sirkulasi udara yang baik. 3. Pemupukan: Hindari pemupukan nitrogen berlebih yang mendorong pertumbuhan jaringan lunak yang rentan infeksi.\nPencegahan: Gunakan benih bebas penyakit, lakukan perlakuan benih dengan fungisida, tanam varietas toleran, dan jaga kebersihan lahan dari sisa tanaman yang terinfeksi.',
    'downy_mildew': '[EMBUN BULU (downy_mildew)]\nGejala: Gejala pada daun berupa lapisan serbuk berwarna putih seperti tepung pada permukaan bawah daun, terutama terlihat pada pagi hari saat embun. Daun yang terinfeksi dapat menunjukkan klorosis (menguning), distorsi, dan pertumbuhan abnormal. Pada serangan sistemik, tanaman menunjukkan gejala "crazy top" di mana daun-daun atas berubah menjadi struktur seperti daun sempit, berkerut, dan malai berubah menjadi daun hijau (phyllody). Tanaman yang terinfeksi berat tampak kerdil dan tidak produktif.\nKondisi Penyebaran: Gejala pada daun berupa lapisan serbuk berwarna putih seperti tepung pada permukaan bawah daun, terutama terlihat pada pagi hari saat embun. Daun yang terinfeksi dapat menunjukkan klorosis (menguning), distorsi, dan pertumbuhan abnormal. Pada serangan sistemik, tanaman menunjukkan gejala "crazy top" di mana daun-daun atas berubah menjadi struktur seperti daun sempit, berkerut, dan malai berubah menjadi daun hijau (phyllody). Tanaman yang terinfeksi berat tampak kerdil dan tidak produktif.\nPenanganan: 1. Pengendalian Kimiawi: Aplikasikan fungisida sistemik berbahan aktif metalaxil (Ridomil Gold 35 WS dosis 2 g/kg benih untuk seed treatment, atau 2 g/L untuk semprot), atau fosetil-Al (Aliette 80 WP dosis 2 g/L). 2. Pengelolaan Air: Kurangi genangan air di persemaian dan pertanaman, perbaiki drainase untuk mengurangi kelembaban berlebih. 3. Sanitasi: Cabut dan musnahkan tanaman yang menunjukkan gejala "crazy top" untuk mencegah penyebaran. 4. Perlakuan Benih: Rendam benih dalam fungisida metalaxil sebelum semai untuk mencegah infeksi dari benih.\nPencegahan: Hindari penggunaan air irigasi dari sumber yang tercemar, perbaiki drainase lahan, gunakan varietas toleran, dan hindari penanaman terlalu rapat yang meningkatkan kelembaban mikro.',
    'healthy': '[healthy]\nGejala: Tanaman sehat tidak menunjukkan gejala penyakit. Ciri-ciri visual yang terdeteksi kamera AI: daun berwarna hijau tua mengkilap tanpa bercak, lesi, atau perubahan warna abnormal. Batang kokoh dan tegak, tidak mudah rebah. Pertumbuhan seragam dan sesuai umur tanaman. Malai berisi penuh dengan gabah bernas pada fase generatif. Tidak ada eksudat, cairan abnormal, atau deformasi pada permukaan tanaman.\nKondisi Penyebaran: Tanaman sehat tidak menunjukkan gejala penyakit. Ciri-ciri visual yang terdeteksi kamera AI: daun berwarna hijau tua mengkilap tanpa bercak, lesi, atau perubahan warna abnormal. Batang kokoh dan tegak, tidak mudah rebah. Pertumbuhan seragam dan sesuai umur tanaman. Malai berisi penuh dengan gabah bernas pada fase generatif. Tidak ada eksudat, cairan abnormal, atau deformasi pada permukaan tanaman.\nPenanganan: 1. Lanjutkan pemupukan berimbang N-P-K sesuai umur dan kebutuhan varietas; tidak diperlukan pestisida apapun. 2. Terapkan sistem irigasi berselang (intermittent irrigation) untuk menghemat air dan mencegah penyakit. 3. Lakukan pemantauan rutin setiap 7-10 hari untuk deteksi dini hama dan penyakit. 4. Pertahankan kebersihan lahan; singkirkan gulma dan sisa tanaman terinfeksi di sekitar petakan. 5. Siapkan jadwal panen sesuai umur varietas dan konfirmasi dengan kondisi visual gabah di lapang.',
    'bacterial_leaf_streak': '[HAWAR DAUN BERGARIS BAKTERI (bacterial_leaf_streak)]\nGejala: Gejala berupa lesi basah berwarna hijau gelap hingga kuning berbentuk garis-garis sempit yang terbatas di antara tulang daun (interveinal). Lesi kemudian berubah warna menjadi coklat kekuningan dan dapat menyatu membentuk area nekrotik yang lebih luas. Pada kondisi lembab, eksudat bakteri berbentuk butiran kecil kekuningan terlihat pada permukaan daun. Gejala dimulai dari daun bagian bawah dan menyebar ke daun atas.\nKondisi Penyebaran: Berkembang optimal pada suhu 25-28°C, kelembaban tinggi >80%, dan kondisi hujan gerimis berkepanjangan. Penularan melalui percikan air hujan, aliran air irigasi, dan kontak langsung antar tanaman.\nPenanganan: 1. Pengendalian Kimiawi: Gunakan bakterisida berbahan aktif bismerthiazol seperti Xanthocide 20 WP dosis 2 g/L air, atau copper-based bactericide seperti Cupravit OB 21 dosis 2-3 g/L air. Interval penyemprotan 7-10 hari. 2. Pengendalian Budidaya: Hindari pemupukan nitrogen berlebihan, perbaiki sirkulasi udara dengan mengatur jarak tanam, kurangi kelembaban mikroklimat tanaman. 3. Drainase: Perbaiki sistem drainase untuk mengurangi kelembaban berlebih di area pertanaman. 4. Rotasi tanaman: Lakukan rotasi dengan tanaman bukan padi untuk memutus siklus patogen.\nPencegahan: Gunakan varietas tahan, lakukan perlakuan benih dengan bakterisida, hindari lahan yang pernah terserang tanpa disanitasi, dan jaga kebersihan alat pertanian untuk mencegah penularan antar petak.',
    'harvest_stage': '[Suhu Udara Tinggi (harvest_stage)]\nGejala: Fase panen bukan merupakan penyakit. Ciri-ciri yang terdeteksi kamera AI: 80–95% gabah pada malai telah berwarna kuning keemasan, malai menunduk akibat bobot gabah yang penuh, daun bendera dan daun atas mulai mengering dan berwarna kuning hingga coklat muda. Kadar air gabah berada pada kisaran 20–25%. Tidak terdapat lesi atau bercak penyakit yang mendominasi — fokus identifikasi adalah kematangan gabah, bukan penyakit aktif.\nKondisi Penyebaran: Rasio gabah bernas terhadap total gabah per malai ≥85%.\nPenanganan: 1. Ukur kadar air gabah dengan moisture meter; jika sudah 20–25%, jadwalkan panen dalam 3–5 hari ke depan saat cuaca cerah. 2. Panen pada pagi hari (06.00–10.00) atau sore hari (15.00–17.00) untuk mengurangi kerontokan dan panas berlebih. 3. Rontokkan gabah dalam 24 jam setelah panen menggunakan power thresher 500–600 rpm untuk mencegah fermentasi. 4. Keringkan gabah hingga kadar air ≤14%; gunakan mechanical dryer 40–43°C jika cuaca tidak mendukung penjemuran alami. 5. Simpan dalam karung bersih atau silo dengan kelembaban gudang <70% dan suhu <25°C untuk mencegah aflatoksin.',
}


def main():
    parser = argparse.ArgumentParser(
        description="Buat Excel penilaian faithfulness (v4 — Penilaian Ground Truth oleh Pakar)"
    )
    parser.add_argument("--input",   default=JSON_PATH, help="Path JSON hasil evaluasi")
    parser.add_argument("--kb",      default=KB_PATH,   help="Path knowledge base .txt")
    parser.add_argument("--output",  default=OUTPUT,    help="Path output Excel")
    parser.add_argument("--dataset", default=DATASET_LABEL, help="Label dataset (mis. A / B)")
    args = parser.parse_args()

    if not Path(args.input).exists():
        print(f"❌ File tidak ditemukan: {args.input}")
        print(f"   Jalankan dulu: python evaluate_rag.py --k 3 --llm all")
        return

    print(f"Membaca JSON    : {args.input}")
    d = load_json(args.input)

    kb_data = {}
    if Path(args.kb).exists():
        print(f"Membaca KB      : {args.kb}")
        kb_data = parse_kb(args.kb)
        print(f"GT-2 ditemukan  : {len(kb_data)} penyakit dari knowledge base")
    else:
        print(f"⚠️  KB tidak ditemukan ({args.kb}), GT-2 akan kosong")

    # Sumber valid: GT-2 diambil persis dari Excel petani (disematkan di GROUND_TRUTH_2),
    # bukan hasil parse ulang KB, supaya dijamin identik.
    kb_data = dict(GROUND_TRUTH_2)
    print(f"GT-2 (disematkan): {len(kb_data)} penyakit dari Excel petani")
    build_excel(d, kb_data, args.output, dataset_label=args.dataset)


if __name__ == "__main__":
    main()
